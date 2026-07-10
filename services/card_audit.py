"""Card Audit service (Phase 6 / Variant A).

End-of-month workflow:
  1. Bookkeeper exports CSV from card provider (Mercury / Revolut / Stripe / generic)
  2. Uploads to FIO via /api/card-audit/import
  3. Service auto-detects format, normalises rows, persists into `card_transactions`
  4. Reconciler matches each card-tx against approved invoices in `documents`:
        - amount equal (±EUR 0.01) on EUR-normalised side
        - posted_at within ±RECONCILE_DAY_WINDOW days of invoice date
        - vendor / counterparty fuzzy match (trigram overlap)
        - invoice.payment_method = 'card' (preferred)
  5. UI shows three buckets: matched / suggested (review) / unmatched (no invoice → policy gap)

This is the manual-CSV bridge while real bank connectors (fio-cashflow Sprint 2-5)
are in backlog. Same DB schema + reconciler approach so the eventual swap to
live API feed is a single endpoint swap, not a data-model change.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import sqlite3
import uuid
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

import config
from services import db
from services import fx

logger = logging.getLogger(__name__)

__all__ = [
    "import_csv",
    "reconcile_period",
    "list_card_tx",
    "get_card_tx",
    "update_card_tx",
    "delete_card_tx",
    "audit_summary",
    "detect_format",
]

# ─────────────────────────────────────────────────────────────────
# Format detection — sniff CSV headers to pick a normaliser
# ─────────────────────────────────────────────────────────────────

# Each format declares: header signature + column→canonical mapping + sign rule.
_FORMAT_SPECS: List[Dict[str, Any]] = [
    {
        # 2026-06-24 FB-C — tightened signature. Was {date, description, amount}
        # which is so generic that EVERY CSV detected as Mercury (false positives
        # filled Past Statement Imports archive with "Mercury" labels for files
        # that were Revolut/etc). Now Mercury REQUIRES Mercury-specific headers
        # (Reference / Bank Description / Account Number) AND text signature.
        "id": "mercury",
        "label": "Mercury Bank",
        "signature": {"date", "amount", "bank description"},
        "extra_signature_any": {"mercury", "mercury technologies"},
        "fields": {
            "date":        ["date", "transaction date", "posting date"],
            "description": ["description", "memo", "name"],
            "amount":      ["amount"],
            "currency":    ["currency"],
            "counterparty":["merchant", "counterparty", "name"],
            "reference":   ["reference", "id"],
        },
        "default_currency": "USD",
        "amount_sign": "natural",  # Mercury uses negative for outflows
    },
    {
        "id": "revolut",
        "label": "Revolut Business",
        # 2026-06-09 — fixed signature to match the ACTUAL Revolut Business
        # CSV export. The earlier spec looked for "started date" / "completed
        # date" but the file uses "Date started (UTC)" / "Date completed
        # (UTC)" and "Payment currency" instead of bare "currency", so the
        # detector fell through to generic and rows didn't parse.
        "signature": {"type", "state", "amount"},
        "extra_signature_any": {"date started (utc)", "date completed (utc)",
                                "payer", "exchange rate", "beneficiary iban"},
        "fields": {
            "date":        ["date completed (utc)", "date started (utc)",
                            "completed date", "started date", "date"],
            "description": ["description", "reference", "payer", "merchant"],
            "amount":      ["amount", "total amount"],
            "currency":    ["payment currency", "currency", "orig currency"],
            "counterparty":["payer", "description"],
            "reference":   ["reference", "id"],
        },
        "default_currency": "EUR",
        "amount_sign": "natural",
    },
    {
        # 2026-06-24 FB-B — Revolut Personal / Retail export (different from Business).
        # Real user CSV had headers: Completed date, Time completed, Status, Transaction
        # type, Counterparty name, Counterparty BIC, Counterparty IBAN, Amount, ...
        "id": "revolut_personal",
        "label": "Revolut Personal",
        "signature": {"completed date", "status", "amount"},
        "extra_signature_any": {"transaction type", "counterparty bic", "counterparty iban",
                                "counterparty name", "time completed", "started date"},
        "fields": {
            "date":        ["completed date", "started date", "date"],
            "description": ["counterparty name", "description", "reference"],
            "amount":      ["amount"],
            "currency":    ["currency", "payment currency"],
            "counterparty":["counterparty name", "merchant", "description"],
            "reference":   ["counterparty iban", "counterparty bic", "reference"],
        },
        "default_currency": "EUR",
        "amount_sign": "natural",
    },
    {
        # 2026-06-30 — Finom Business statement. Same column family as
        # Revolut Personal (Completed date / Time completed / Status /
        # Transaction type / Counterparty name / BIC / IBAN / Reference)
        # but uses "Payment amount" + "Payment currency" instead of bare
        # "Amount", and has wallet-specific columns ("Wallet name",
        # "Wallet IBAN", "Wallet balance after transaction"). Without
        # this spec, Finom CSVs fall through to Generic CSV and the
        # importer can't find a date column it understands.
        # Real export the operator pasted on 2026-06-30:
        #   27.06.2026,09:53,Completed,Card,FACEBK *5MDQ5VMQG2,,…
        "id": "finom",
        "label": "Finom Business",
        "signature": {"completed date", "status", "payment amount"},
        "extra_signature_any": {"wallet name", "wallet iban",
                                "wallet balance after transaction",
                                "payment currency", "transaction id"},
        "fields": {
            "date":        ["completed date", "date"],
            "description": ["counterparty name", "reference", "description"],
            "amount":      ["payment amount", "original amount", "amount"],
            "currency":    ["payment currency", "original currency", "currency"],
            "counterparty":["counterparty name", "transaction payer"],
            "reference":   ["reference", "counterparty iban", "transaction id"],
        },
        "default_currency": "EUR",
        "amount_sign": "natural",
    },
    {
        # 2026-06-24 FB-C — Stripe signature was {description, amount, currency}
        # which collided with generic bank CSVs. Now requires a Stripe-specific
        # column to avoid false positives.
        "id": "stripe",
        "label": "Stripe",
        "signature": {"amount", "currency", "balance transaction id"},
        "extra_signature_any": {"stripe", "fee", "payment intent id"},
        "fields": {
            "date":        ["created", "available on", "date"],
            "description": ["description"],
            "amount":      ["amount", "net"],
            "currency":    ["currency"],
            "counterparty":["description", "customer"],
            "reference":   ["balance transaction id", "transaction id", "id"],
        },
        "default_currency": "USD",
        "amount_sign": "natural",
    },
    {
        "id": "airwallex",
        "label": "Airwallex",
        "signature": {"transaction date", "amount"},
        "extra_signature_any": {"airwallex", "wallet balance"},
        "fields": {
            "date":        ["transaction date", "value date", "date"],
            "description": ["description", "merchant name"],
            "amount":      ["amount", "transaction amount"],
            "currency":    ["currency", "transaction currency"],
            "counterparty":["merchant name", "description"],
            "reference":   ["reference", "id"],
        },
        "default_currency": "USD",
        "amount_sign": "natural",
    },
    {
        # 2026-07-10 — Narvi Payments (Estonian EMI, narvi.com). Real export
        # header (operator statement 2026-06): Transaction Id, Transaction
        # date, Transaction type, Currency, Transaction amount, Fee Amount,
        # Net credited amount, Net debited amount, Transaction description,
        # Sender IBAN/BIC/name/address…, Recipient IBAN/BIC/name/address…
        # Before this spec the file fell through to Generic and every row was
        # dropped as "no_amount" (Narvi's money column is "Transaction amount",
        # not bare "amount"), i.e. a 0-row import that stalled month-close.
        # "Transaction amount" is ALREADY signed: Credit is +, Debit/Fee is -.
        "id": "narvi",
        "label": "Narvi Payments",
        "signature": {"transaction date", "transaction amount", "transaction type"},
        "extra_signature_any": {"net credited amount", "net debited amount",
                                "sender iban", "recipient iban", "transaction id",
                                "narvi"},
        "fields": {
            "date":        ["transaction date", "date"],
            "description": ["transaction description", "description", "reference"],
            "amount":      ["transaction amount", "amount"],
            "currency":    ["currency"],
            "counterparty":["recipient name", "sender name", "transaction description"],
            "reference":   ["transaction id", "reference"],
        },
        # For a money-IN row the counterparty is the SENDER; for money-OUT it
        # is the RECIPIENT (the holder sits on the other side). Picking by sign
        # avoids labelling an incoming invoice payment with our own name.
        "counterparty_by_sign": {"in": ["sender name"], "out": ["recipient name"]},
        "default_currency": "EUR",
        "amount_sign": "natural",
    },
    {
        # 2026-07-10 — PayPal activity CSV (the "CSR" download from
        # paypal.com → Reports → Activity). Real header: Date, Time, Time
        # Zone, Description, Currency, Gross, Fee, Net, Balance, Transaction
        # ID, From Email Address, Name, Bank Name, Bank account, Postage…,
        # VAT, Invoice ID, Reference Txn ID. One file mixes currencies
        # (EUR + GBP rows) — per-row `Currency` drives FX. Before this spec
        # PayPal fell through to Generic and every row dropped as "no_amount"
        # (the value column is "Gross"/"Net", never bare "amount").
        # "Gross" is the transaction value (already signed: outflow -, inflow +);
        # "Net" = Gross minus PayPal's fee. We match on Gross (= invoice value).
        "id": "paypal",
        "label": "PayPal",
        "signature": {"gross", "net", "transaction id"},
        "extra_signature_any": {"from email address", "time zone", "balance",
                                "invoice id", "reference txn id", "paypal"},
        "fields": {
            "date":        ["date"],
            "description": ["description", "name"],
            "amount":      ["gross", "net", "amount"],
            "currency":    ["currency"],
            "counterparty":["name", "from email address", "description"],
            "reference":   ["transaction id", "invoice id"],
        },
        "default_currency": "EUR",
        "amount_sign": "natural",
    },
    {
        "id": "generic",
        "label": "Generic CSV",
        "signature": set(),  # last-resort fallback
        "extra_signature_any": set(),
        "fields": {
            "date":        ["date", "transaction date", "posted at", "posting date", "value date"],
            "description": ["description", "memo", "narrative", "details"],
            "amount":      ["amount", "value", "sum", "debit", "credit"],
            "currency":    ["currency", "ccy"],
            "counterparty":["counterparty", "merchant", "payee", "vendor"],
            "reference":   ["reference", "id", "transaction id"],
        },
        "default_currency": "EUR",
        "amount_sign": "natural",
    },
]


def detect_format(headers: List[str]) -> Dict[str, Any]:
    """Pick best-matching format spec by header overlap."""
    norm_headers = {h.strip().lower() for h in headers if h}
    best = _FORMAT_SPECS[-1]  # generic fallback
    best_score = -1
    for spec in _FORMAT_SPECS:
        sig = spec.get("signature", set())
        extras = spec.get("extra_signature_any", set())
        if sig and not sig.issubset(norm_headers):
            continue
        score = len(sig & norm_headers)
        if any(e in " ".join(norm_headers) for e in extras):
            score += 2
        if score > best_score:
            best_score = score
            best = spec
    return best


# ─────────────────────────────────────────────────────────────────
# Account → legal-entity / profit-center attribution
# ─────────────────────────────────────────────────────────────────
# Which company does this statement belong to? We answer it by scanning the
# raw statement text for an IBAN / account number / BIC / holder-name token
# registered in data/bank_accounts.json. This is what lets upload be
# "automatic": the operator no longer has to remember that the Narvi export
# is Amitours Holding and the Airwallex export is Amitours Group SA.

_BANK_ACCOUNTS_CACHE: Dict[str, Any] = {"mtime": None, "accounts": []}


def _bank_accounts_path() -> str:
    import os
    return os.path.join(os.path.dirname(config.DB_PATH), "bank_accounts.json")


def _load_bank_accounts() -> List[Dict[str, Any]]:
    """Load + cache the account registry (mtime-invalidated so admin edits
    to data/bank_accounts.json take effect without a restart)."""
    import os
    path = _bank_accounts_path()
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        return []
    if _BANK_ACCOUNTS_CACHE["mtime"] == mtime:
        return _BANK_ACCOUNTS_CACHE["accounts"]
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        accounts = data.get("accounts") or []
    except (OSError, json.JSONDecodeError, AttributeError) as exc:
        logger.warning("bank_accounts.json load failed: %s", exc)
        accounts = []
    _BANK_ACCOUNTS_CACHE["mtime"] = mtime
    _BANK_ACCOUNTS_CACHE["accounts"] = accounts
    return accounts


def detect_account(raw_text: str) -> Optional[Dict[str, Any]]:
    """Scan statement text for a registered account token. First match wins
    (registry is ordered most-specific-first). Returns
    {id, label, legal_entity, profit_center, matched_on} or None.

    Match is a whitespace-insensitive, case-insensitive substring test so an
    IBAN printed with spaces ("GB51 REVO 0099…") still matches "GB51REVO0099…".
    """
    if not raw_text:
        return None
    haystack = re.sub(r"\s+", "", raw_text).lower()
    for acct in _load_bank_accounts():
        for token in acct.get("match", []):
            if not token:
                continue
            needle = re.sub(r"\s+", "", str(token)).lower()
            if needle and needle in haystack:
                return {
                    "id":            acct.get("id"),
                    "label":         acct.get("label"),
                    "legal_entity":  acct.get("legal_entity"),
                    "profit_center": acct.get("profit_center"),
                    "matched_on":    token,
                }
    return None


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

def _pick_column(row: Dict[str, Any], candidates: List[str]) -> Optional[str]:
    """Return first non-empty value among candidate column names."""
    lo = {k.strip().lower(): v for k, v in row.items() if k}
    for c in candidates:
        v = lo.get(c)
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return None


def _parse_amount(s: Optional[str]) -> Optional[float]:
    # 2026-07-08 (C8) -- delegate to the shared money parser which
    # correctly distinguishes US/EU thousands vs decimal separators.
    # The old `s.replace(",", ".")` turned "1,500" into 1.5 (a EUR 1,500
    # card charge imported as EUR 1.50). See services/money.parse_money.
    from services.money import parse_money
    return parse_money(s)


# Month-name lookup so "30 Jun 2026" / "Jun 16 2025" parse without locale deps.
_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}


def _parse_date(s: Optional[str]) -> Optional[str]:
    """Return ISO YYYY-MM-DD or None. Bank exports are a zoo of formats, so
    this is deliberately forgiving:
      • ISO with a time/zone suffix   2026-06-27T09:53:00Z / +03:00
      • EU / US numeric               27.06.2026 · 27/06/2026 · 06/27/2026
      • month-name, either order      30 Jun 2026 EEST · Jun 16 2025
      • 2-digit years                 27/06/26
    A row whose date won't parse is dropped as `no_date`, so a too-strict
    parser silently eats real transactions — hence the wide net."""
    if not s:
        return None
    s = str(s).strip()

    # 1. ISO prefix (optionally followed by T/space + time + tz). Most banks.
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    # 2. Month-name forms: "30 Jun 2026", "Jun 16 2025", "30-Jun-2026".
    mn = re.search(r"(?:(\d{1,2})[ /\-]+([A-Za-z]{3,})|([A-Za-z]{3,})[ /\-]+(\d{1,2}))[ /,\-]+(\d{4})", s)
    if mn:
        if mn.group(1):
            day, mon_name, year = mn.group(1), mn.group(2), mn.group(5)
        else:
            day, mon_name, year = mn.group(4), mn.group(3), mn.group(5)
        mon = _MONTHS.get(mon_name[:3].lower())
        if mon:
            try:
                return datetime(int(year), mon, int(day)).strftime("%Y-%m-%d")
            except ValueError:
                pass

    # 3. Numeric d/m/y or m/d/y or d.m.y, 2- or 4-digit year. Take the
    #    leading date token only (ignore any trailing time).
    tok = re.match(r"^(\d{1,2})[./\-](\d{1,2})[./\-](\d{2,4})", s)
    if tok:
        a, b, y = int(tok.group(1)), int(tok.group(2)), int(tok.group(3))
        if y < 100:
            y += 2000
        # Disambiguate day vs month: if the first field can't be a month
        # (>12) it's a day (EU order); else assume EU (d/m/y) which is what
        # every non-US bank in this holding uses (Narvi/Revolut/Finom/PayPal).
        if a > 12 and b <= 12:
            day, mon = a, b
        elif b > 12 and a <= 12:
            day, mon = b, a           # US m/d/y (e.g. Mercury)
        else:
            day, mon = a, b           # ambiguous → EU default
        try:
            return datetime(y, mon, day).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def _make_id(source: str, batch_id: str, row: Dict[str, Any]) -> str:
    # 2026-07-08 (H1) — fold in the row's position within its import file
    # (row_seq). Two genuinely-identical charges in one statement
    # (2x EUR 4.50 Starbucks same day) now get DISTINCT ids and both
    # survive; re-importing the same file yields the same row_seq -> same
    # ids -> PK conflict -> idempotent skip. Excludes batch_id so re-import
    # dedups across batches.
    seq = row.get("row_seq", 0)
    payload = (f"{source}|{row.get('posted_at','')}|{row.get('amount','')}|"
               f"{row.get('description','')}|{seq}").lower()
    h = hashlib.sha1(payload.encode("utf-8")).hexdigest()[:14]
    return f"ct_{h}"


# ─────────────────────────────────────────────────────────────────
# Import
# ─────────────────────────────────────────────────────────────────

def import_statement(
    file_bytes: bytes,
    filename: str,
    imported_by: str = "user",
    source_override: Optional[str] = None,
    profit_center: Optional[str] = None,
) -> Dict[str, Any]:
    """Dispatch a bank statement file (CSV / XLSX / PDF) into the CSV pipeline.

    XLSX and PDF are converted to CSV-shaped bytes and routed through the
    existing import_csv() parser. Each row in the resulting batch can be
    tagged with `profit_center` so the stakeholder's own stream is pre-set.
    """
    name_low = (filename or "").lower()
    # Text used ONLY for account (legal-entity) detection — we want the full
    # original content incl. IBANs/holder, which the PDF→CSV reduction drops.
    account_text = ""
    if name_low.endswith(".xlsx") or name_low.endswith(".xls"):
        try:
            csv_bytes = _xlsx_to_csv_bytes(file_bytes)
        except Exception as exc:
            raise ValueError("Failed to parse XLSX: %s" % exc)
        account_text = csv_bytes.decode("utf-8", "ignore")
    elif name_low.endswith(".pdf"):
        try:
            account_text = _pdf_extract_text(file_bytes)
        except Exception:  # noqa: BLE001 — detection is best-effort; never block import
            account_text = ""
        try:
            csv_bytes = _pdf_to_csv_bytes(file_bytes)
        except Exception as exc:
            raise ValueError("Failed to parse PDF: %s" % exc)
    else:
        csv_bytes = file_bytes  # assume CSV / TSV
        try:
            account_text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            account_text = file_bytes.decode("latin-1", "ignore")

    # Which company owns this statement? (Narvi→Holding, Airwallex→Group SA…)
    account = detect_account(account_text)
    legal_entity = account.get("legal_entity") if account else None

    result = import_csv(csv_bytes, filename, imported_by=imported_by,
                        source_override=source_override,
                        legal_entity=legal_entity)

    # Profit-center precedence: explicit operator choice > registry default
    # (only set for single-stream accounts). Multi-stream accounts leave it
    # to per-row vendor heuristics, so we never blanket-stamp a wrong PC.
    pc = None
    if profit_center and profit_center.strip():
        pc = profit_center.strip().upper()[:4]
        result["profit_center_source"] = "operator"
    elif account and account.get("profit_center"):
        pc = str(account["profit_center"]).strip().upper()[:4]
        result["profit_center_source"] = "account_registry"
    if pc:
        conn = db.get_connection()
        try:
            conn.execute(
                "UPDATE card_transactions SET profit_center = ? WHERE batch_id = ?",
                (pc, result["batch_id"]),
            )
            conn.commit()
        finally:
            conn.close()
        result["profit_center_stamped"] = pc

    # Surface the detected account so the UI can confirm / correct the entity.
    result["account"] = account
    result["legal_entity"] = legal_entity
    # A PDF whose layout the line-heuristic can't read yields 0 rows. Say so
    # explicitly (and name the detected entity) so the operator knows to grab
    # the CSV export instead of staring at a silent success.
    if name_low.endswith(".pdf") and result.get("inserted", 0) == 0:
        result["diagnosis"] = "pdf_layout_unsupported"
        ent = (account or {}).get("label") or "this account"
        result["hint"] = (
            "This PDF's transaction layout couldn't be read automatically. "
            f"Recognised the account as {ent}. Please upload the CSV/XLSX "
            "export of the same statement — every bank here (Narvi, Revolut, "
            "Finom, PayPal, Mercury, Airwallex) offers one, and CSV imports "
            "cleanly."
        )
    return result


def _xlsx_to_csv_bytes(xlsx_bytes: bytes) -> bytes:
    """Convert XLSX → CSV bytes. First worksheet only, all cells stringified."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb.active
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if c is None else str(c) for c in row])
    wb.close()
    return buf.getvalue().encode("utf-8")


# Bank-statement PDF parsing heuristic.
# Lines of interest typically look like:
#   "01.05.2026   PAYMENT TO ACME LTD   -1,234.56 EUR"
#   "2026-05-01   ACME LTD   EUR  1,234.56"
# We extract any line containing a date AND a number that looks like an amount.
_PDF_DATE_RE = re.compile(
    r"(?P<date>(?:\d{1,2}[./-]\d{1,2}[./-]\d{2,4})|(?:\d{4}[./-]\d{1,2}[./-]\d{1,2}))"
)
_PDF_AMOUNT_RE = re.compile(
    r"(?P<sign>[-+]?)\s*"
    r"(?P<num>\d{1,3}(?:[ ,.]\d{3})*[.,]\d{2}|\d+[.,]\d{2})"
    r"\s*(?P<ccy>EUR|USD|GBP|CHF|PLN|SEK|NOK|DKK|€|\$|£)?"
)


def _pdf_extract_text(pdf_bytes: bytes) -> str:
    """Return the full text of a PDF (all pages joined). One damaged page is
    swallowed rather than aborting the whole extract. Shared by account
    detection (needs the IBAN/holder) and the transaction heuristic."""
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(pdf_bytes))
    parts = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:  # noqa: BLE001 — pypdf raises many types on damaged pages
            parts.append("")
    return "\n".join(parts)


def _pdf_to_csv_bytes(pdf_bytes: bytes) -> bytes:
    """Extract transactions from a bank-statement PDF into CSV bytes.

    Heuristic: any line that contains BOTH a date pattern and an amount
    pattern is treated as a transaction. Description = whatever text sits
    between the date and the amount.

    Output columns match the 'generic' format spec so detect_format() can
    consume it: posted_at, amount, currency, description.

    NOTE: this line-based heuristic handles one-transaction-per-line PDFs.
    Statements that split a transaction across several lines (Mercury,
    Airwallex, Narvi PDF) may yield few/zero rows — in that case the CSV
    export is the reliable path, and import_csv() emits a loud
    diagnosis="no_rows_parsed" rather than a silent success. Account
    (legal-entity) detection still works on such PDFs because it runs on the
    full extracted text, not this reduced CSV.
    """
    text = _pdf_extract_text(pdf_bytes)

    rows: List[Tuple[str, str, str, str]] = []  # (date, amount, currency, description)
    for raw_line in text.split("\n"):
        line = raw_line.strip()
        if len(line) < 8:
            continue
        d = _PDF_DATE_RE.search(line)
        if not d:
            continue
        a = _PDF_AMOUNT_RE.search(line[d.end():]) or _PDF_AMOUNT_RE.search(line)
        if not a:
            continue
        date_s = d.group("date")
        amt_raw = a.group("num").replace(" ", "")
        # European format "1.234,56" → "1234.56"; US format "1,234.56" already OK
        if amt_raw.count(",") == 1 and amt_raw.count(".") >= 1:
            amt_raw = amt_raw.replace(".", "").replace(",", ".")
        elif amt_raw.count(",") == 1 and amt_raw.count(".") == 0:
            amt_raw = amt_raw.replace(",", ".")
        elif amt_raw.count(",") > 1:
            amt_raw = amt_raw.replace(",", "")
        sign = a.group("sign") or ""
        amt = sign + amt_raw
        ccy_raw = (a.group("ccy") or "").upper()
        ccy = {"€": "EUR", "$": "USD", "£": "GBP"}.get(ccy_raw, ccy_raw) or "EUR"

        # Description = text between date end and amount start
        desc = line[d.end():a.start() if a.re == _PDF_AMOUNT_RE and a is not None else None].strip()
        if not desc or len(desc) < 2:
            desc = line[:d.start()].strip() or line.strip()
        # Trim trailing tokens (running balance numbers, etc.)
        desc = re.sub(r"\s+", " ", desc)[:200]

        rows.append((date_s, amt, ccy, desc))

    if not rows:
        # If we found nothing, surface a synthetic "no rows" CSV that detect_format
        # will still accept (zero data rows).
        return b"date,amount,currency,description\n"

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["date", "amount", "currency", "description"])
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def import_csv(
    csv_bytes: bytes,
    filename: str,
    imported_by: str = "user",
    source_override: Optional[str] = None,
    legal_entity: Optional[str] = None,
) -> Dict[str, Any]:
    """Parse a CSV statement, persist into card_transactions, dedupe by UNIQUE key.

    Returns:
      {batch_id, source, total_rows, inserted, skipped, errors, sample}
    """
    try:
        text = csv_bytes.decode("utf-8-sig")  # handles BOM
    except UnicodeDecodeError:
        text = csv_bytes.decode("latin-1")

    # Auto-detect delimiter
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # default to comma
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = reader.fieldnames or []
    spec = _FORMAT_SPECS[-1]
    if source_override:
        spec = next((s for s in _FORMAT_SPECS if s["id"] == source_override), spec)
    else:
        spec = detect_format(headers)

    batch_id = "b_" + uuid.uuid4().hex[:10]
    now_iso = datetime.utcnow().isoformat()
    inserted = 0
    skipped = 0
    errors: List[Dict[str, Any]] = []
    sample_rows: List[Dict[str, Any]] = []
    # 2026-06-24 FB-2a — track WHY rows are skipped so the UI can show
    # "your CSV format wasn't recognised" instead of silently importing 0 rows
    skip_reasons: Dict[str, int] = {"no_date": 0, "no_amount": 0, "zero_amount": 0,
                                     "duplicate": 0, "other": 0}
    # 2026-07-10 — keep a few concrete skipped rows (not just tallies) so the
    # UI can show the operator EXACTLY which line + which columns failed,
    # instead of a silent "0 imported". Cap at 5 to bound the payload.
    skipped_samples: List[Dict[str, Any]] = []
    def _note_skip(reason: str, ridx: int, raw: Dict[str, Any]) -> None:
        if len(skipped_samples) < 5:
            skipped_samples.append({
                "row": ridx, "reason": reason,
                "date_seen": _pick_column(raw, spec["fields"]["date"]),
                "amount_seen": _pick_column(raw, spec["fields"]["amount"]),
            })

    conn = db.get_connection()
    try:
        for row_idx, raw_row in enumerate(reader, start=2):  # row_idx 2 = first data row
            try:
                date_s = _pick_column(raw_row, spec["fields"]["date"])
                amount_s = _pick_column(raw_row, spec["fields"]["amount"])
                desc = _pick_column(raw_row, spec["fields"]["description"]) or ""
                ccy = _pick_column(raw_row, spec["fields"]["currency"]) or spec["default_currency"]
                counterparty = _pick_column(raw_row, spec["fields"]["counterparty"]) or desc
                reference = _pick_column(raw_row, spec["fields"]["reference"]) or ""

                posted = _parse_date(date_s)
                amount = _parse_amount(amount_s)
                if posted is None:
                    skipped += 1; skip_reasons["no_date"] += 1; _note_skip("no_date", row_idx, raw_row); continue
                if amount is None:
                    skipped += 1; skip_reasons["no_amount"] += 1; _note_skip("no_amount", row_idx, raw_row); continue
                if amount == 0:
                    skipped += 1; skip_reasons["zero_amount"] += 1; continue

                # Sign-aware counterparty (e.g. Narvi: inflow→sender, outflow→recipient)
                cbs = spec.get("counterparty_by_sign")
                if cbs:
                    cols = cbs.get("in" if amount > 0 else "out") or []
                    cp = _pick_column(raw_row, cols)
                    if cp:
                        counterparty = cp

                # FX → EUR
                fx_data = fx.convert_to_eur(amount, ccy.upper(), posted)

                row = {
                    "source":         spec["id"],
                    "batch_id":       batch_id,
                    "imported_at":    now_iso,
                    "imported_by":    imported_by,
                    "posted_at":      posted,
                    "period":         posted[:7],
                    "amount":         round(amount, 2),
                    "currency":       ccy.upper()[:3],
                    "amount_eur":     fx_data["amount_eur"],
                    "fx_rate":        fx_data["fx_rate"],
                    "fx_date":        fx_data["fx_date"],
                    "description":    desc[:500],
                    "counterparty":   counterparty[:200],
                    "reference":      reference[:120],
                    "card_holder":    None,
                    "department":     None,
                    "profit_center":  None,
                    "legal_entity":   legal_entity,
                    "matched_invoice_id": None,
                    "match_status":   "unmatched",
                    "match_confidence": 0,
                    "match_reason":   None,
                    "notes":          None,
                    "raw_row":        json.dumps(raw_row, ensure_ascii=False)[:2000],
                    "row_seq":        row_idx,   # H1: file position → distinct id for genuine dupes
                }
                row["id"] = _make_id(spec["id"], batch_id, row)

                try:
                    conn.execute(
                        """INSERT INTO card_transactions
                           (id, source, batch_id, imported_at, imported_by, posted_at,
                            period, amount, currency, amount_eur, fx_rate, fx_date,
                            description, counterparty, reference, card_holder,
                            department, profit_center, legal_entity, matched_invoice_id,
                            match_status, match_confidence, match_reason, notes, raw_row,
                            row_seq)
                           VALUES (:id, :source, :batch_id, :imported_at, :imported_by,
                                   :posted_at, :period, :amount, :currency, :amount_eur,
                                   :fx_rate, :fx_date, :description, :counterparty,
                                   :reference, :card_holder, :department, :profit_center,
                                   :legal_entity, :matched_invoice_id, :match_status, :match_confidence,
                                   :match_reason, :notes, :raw_row, :row_seq)""",
                        row,
                    )
                    inserted += 1
                    if len(sample_rows) < 3:
                        sample_rows.append({
                            "posted_at": row["posted_at"], "amount": row["amount"],
                            "currency": row["currency"], "amount_eur": row["amount_eur"],
                            "description": row["description"][:60], "counterparty": row["counterparty"][:40],
                        })
                except sqlite3.IntegrityError:
                    skipped += 1; skip_reasons["duplicate"] += 1
            except Exception as exc:  # noqa: BLE001 — per-row defensive boundary
                errors.append({"row": row_idx, "error": str(exc)[:120]})
                skip_reasons["other"] += 1
        conn.commit()
    finally:
        conn.close()

    logger.info(
        "card_audit import %s: source=%s inserted=%d skipped=%d errors=%d",
        filename, spec["id"], inserted, skipped, len(errors),
    )
    # FB-2a — if NOTHING was inserted, build a loud diagnostic so the UI can
    # tell the user exactly why instead of looking like it succeeded silently.
    diagnosis = None
    if inserted == 0:
        if skipped == 0 and not errors:
            diagnosis = "no_rows_parsed"   # empty or completely-broken CSV
        elif skip_reasons["no_date"] == skipped and skipped > 0:
            diagnosis = "date_column_not_found"
        elif skip_reasons["no_amount"] == skipped and skipped > 0:
            diagnosis = "amount_column_not_found"
        elif skip_reasons["duplicate"] == skipped:
            diagnosis = "all_duplicate"
        else:
            diagnosis = "format_mismatch"   # mix of reasons → format spec wrong

    return {
        "batch_id":    batch_id,
        "source":      spec["id"],
        "source_label": spec["label"],
        "filename":    filename,
        "total_rows":  inserted + skipped,
        "inserted":    inserted,
        "skipped":     skipped,
        "skip_reasons": skip_reasons,
        "errors":      errors[:10],
        "sample":      sample_rows,
        "skipped_samples": skipped_samples,
        "headers_seen": headers,
        "diagnosis":   diagnosis,
    }


# ─────────────────────────────────────────────────────────────────
# Reconciler — match card-tx ↔ FIO invoices
# ─────────────────────────────────────────────────────────────────

# Tunables — same defaults as fio-cashflow reconciler for consistency
AMOUNT_TOLERANCE_EUR = 0.01
DAY_WINDOW = 3
AUTO_MATCH_CONFIDENCE = 70
SUGGEST_CONFIDENCE = 40


def _normalize_name(s: Optional[str]) -> str:
    if not s:
        return ""
    s = re.sub(r"\b(sia|as|ltd|gmbh|sarl|ou|oy|inc|llc|b\.v\.|s\.a\.|s\.p\.a\.)\b", "", s.lower())
    return re.sub(r"[^a-z0-9]", "", s)


def _vendor_similarity(a: Optional[str], b: Optional[str]) -> float:
    na, nb = _normalize_name(a), _normalize_name(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    if na in nb or nb in na:
        return 0.85
    def trig(s):
        return {s[i:i + 3] for i in range(len(s) - 2)} if len(s) >= 3 else {s}
    ta, tb = trig(na), trig(nb)
    inter = ta & tb
    union = ta | tb
    return len(inter) / len(union) if union else 0.0


def _list_invoice_candidates(period: str) -> List[Dict[str, Any]]:
    """All approved/posted/classified card-invoices in this period."""
    conn = db.get_connection()
    try:
        # Search ±1 month around period to catch boundary cases
        prefix = period[:7]
        try:
            y, m = int(prefix[:4]), int(prefix[5:7])
            prev = f"{y if m > 1 else y-1:04d}-{(m-1) if m > 1 else 12:02d}"
            nxt  = f"{y if m < 12 else y+1:04d}-{(m+1) if m < 12 else 1:02d}"
        except Exception:
            prev = nxt = prefix
        rows = conn.execute(
            """SELECT id, vendor, amount, currency, period,
                      payment_method, ledger_code, profit_center, department,
                      parsed_json, classification_json
               FROM documents
               WHERE status IN ('classified','approved','posted')
                 AND period IN (?, ?, ?)""",
            (prev, prefix, nxt),
        ).fetchall()
        return [db._row_to_dict(r) for r in rows]
    finally:
        conn.close()


def _doc_date(d: Dict[str, Any]) -> Optional[datetime]:
    """Best document date for matching: parsed_json.dates.document_date, else period+15.

    parsed_json may be a dict (when row came through _row_to_dict) or a raw
    JSON string (legacy callers) — handle both.
    """
    parsed_raw = d.get("parsed_json")
    try:
        if isinstance(parsed_raw, str):
            parsed = json.loads(parsed_raw or "{}")
        elif isinstance(parsed_raw, dict):
            parsed = parsed_raw
        else:
            parsed = {}
        dt = (parsed.get("dates") or {}).get("document_date")
        if dt:
            return datetime.fromisoformat(dt[:10])
    except (ValueError, TypeError, json.JSONDecodeError):
        pass
    if d.get("period"):
        try:
            return datetime.fromisoformat(d["period"] + "-15")
        except Exception:
            return None
    return None


def _score_card_vs_invoice(tx: Dict[str, Any], inv: Dict[str, Any]) -> Tuple[int, str]:
    reasons: List[str] = []
    score = 0

    # 1. Amount on EUR-normalised side (card always converted; invoice.amount is EUR after Phase 2.1)
    tx_amt = abs(float(tx.get("amount_eur") or tx.get("amount") or 0))
    inv_amt = abs(float(inv.get("amount") or 0))
    if not tx_amt or not inv_amt:
        return 0, ""
    if abs(tx_amt - inv_amt) <= AMOUNT_TOLERANCE_EUR:
        score += 50
        reasons.append(f"€{tx_amt:.2f} exact")
    elif abs(tx_amt - inv_amt) / max(tx_amt, inv_amt) < 0.015:
        score += 35
        reasons.append(f"€{tx_amt:.2f}≈€{inv_amt:.2f}")
    else:
        return 0, ""

    # 2. Date proximity
    try:
        tx_date = datetime.fromisoformat(tx["posted_at"][:10])
    except Exception:
        return 0, ""
    inv_date = _doc_date(inv)
    if inv_date:
        delta = abs((tx_date - inv_date).days)
        if delta <= DAY_WINDOW:
            score += 20
            reasons.append(f"date ±{delta}d")
        elif delta <= 14:
            score += 8
            reasons.append(f"date {delta}d apart")

    # 3. Vendor similarity
    sim = _vendor_similarity(tx.get("counterparty") or tx.get("description"), inv.get("vendor"))
    if sim >= 0.85:
        score += 20
        reasons.append(f"vendor {sim:.0%}")
    elif sim >= 0.5:
        score += 10
        reasons.append(f"vendor partial {sim:.0%}")

    # 4. Payment method match — bonus if invoice flagged card
    if (inv.get("payment_method") or "").lower() == "card":
        score += 8
        reasons.append("inv.payment=card")

    return min(score, 100), "; ".join(reasons)


def reconcile_period(period: str) -> Dict[str, int]:
    """Re-run matching for all card-tx in given period. Returns count summary."""
    invoices = _list_invoice_candidates(period)
    conn = db.get_connection()
    counts = {"checked": 0, "auto": 0, "suggested": 0, "unmatched": 0, "stayed": 0}
    try:
        rows = conn.execute(
            "SELECT * FROM card_transactions WHERE period = ? AND match_status != 'manual'",
            (period,),
        ).fetchall()
        for tx in rows:
            tx_d = dict(tx)
            counts["checked"] += 1
            best_score = 0
            best_inv = None
            best_reason = ""
            for inv in invoices:
                s, r = _score_card_vs_invoice(tx_d, inv)
                if s > best_score:
                    best_score, best_inv, best_reason = s, inv, r
            if best_inv and best_score >= AUTO_MATCH_CONFIDENCE:
                conn.execute(
                    "UPDATE card_transactions SET matched_invoice_id=?, match_status='matched', match_confidence=?, match_reason=? WHERE id=?",
                    (best_inv["id"], best_score, best_reason, tx_d["id"]),
                )
                counts["auto"] += 1
            elif best_inv and best_score >= SUGGEST_CONFIDENCE:
                conn.execute(
                    "UPDATE card_transactions SET matched_invoice_id=?, match_status='suggested', match_confidence=?, match_reason=? WHERE id=?",
                    (best_inv["id"], best_score, best_reason, tx_d["id"]),
                )
                counts["suggested"] += 1
            else:
                conn.execute(
                    "UPDATE card_transactions SET matched_invoice_id=NULL, match_status='unmatched', match_confidence=0, match_reason=? WHERE id=?",
                    ("no candidate above threshold", tx_d["id"]),
                )
                counts["unmatched"] += 1
        conn.commit()
    finally:
        conn.close()
    logger.info("reconcile_period %s: %s", period, counts)
    return counts


# ─────────────────────────────────────────────────────────────────
# CRUD + audit summary
# ─────────────────────────────────────────────────────────────────

def list_card_tx(
    period: Optional[str] = None,
    department: Optional[str] = None,
    card_holder: Optional[str] = None,
    match_status: Optional[str] = None,
    limit: int = 500,
) -> List[Dict[str, Any]]:
    conn = db.get_connection()
    where: List[str] = []
    params: List[Any] = []
    if period:
        where.append("period = ?")
        params.append(period)
    if department:
        where.append("department = ?")
        params.append(department)
    if card_holder:
        where.append("card_holder = ?")
        params.append(card_holder)
    if match_status:
        where.append("match_status = ?")
        params.append(match_status)
    sql = "SELECT * FROM card_transactions"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY posted_at DESC LIMIT ?"
    params.append(limit)
    try:
        return [db._row_to_dict(r) for r in conn.execute(sql, params).fetchall()]
    finally:
        conn.close()


def get_card_tx(tx_id: str) -> Optional[Dict[str, Any]]:
    conn = db.get_connection()
    try:
        row = conn.execute("SELECT * FROM card_transactions WHERE id=?", (tx_id,)).fetchone()
        return db._row_to_dict(row) if row else None
    finally:
        conn.close()


def update_card_tx(tx_id: str, fields: Dict[str, Any]) -> None:
    if not fields:
        return
    set_clause = ", ".join(f"{k} = ?" for k in fields)
    conn = db.get_connection()
    try:
        conn.execute(
            f"UPDATE card_transactions SET {set_clause} WHERE id = ?",
            list(fields.values()) + [tx_id],
        )
        conn.commit()
    finally:
        conn.close()


def delete_card_tx(tx_id: str) -> None:
    conn = db.get_connection()
    try:
        conn.execute("DELETE FROM card_transactions WHERE id=?", (tx_id,))
        conn.commit()
    finally:
        conn.close()


def audit_summary(period: str) -> Dict[str, Any]:
    """Top-level numbers for the Card Audit dashboard."""
    conn = db.get_connection()
    try:
        rows = conn.execute(
            """SELECT match_status, department, card_holder,
                      COUNT(*) AS n,
                      COALESCE(SUM(ABS(amount_eur)),0) AS total_abs_eur
               FROM card_transactions
               WHERE period = ?
               GROUP BY match_status, department, card_holder""",
            (period,),
        ).fetchall()
    finally:
        conn.close()

    by_status: Dict[str, Dict[str, float]] = {}
    by_department: Dict[str, Dict[str, Any]] = {}
    by_holder: Dict[str, Dict[str, Any]] = {}
    total = {"n": 0, "eur": 0.0}
    for r in rows:
        st = r["match_status"]
        dept = r["department"] or "(unassigned)"
        holder = r["card_holder"] or "(unassigned)"
        by_status.setdefault(st, {"n": 0, "eur": 0.0})
        by_status[st]["n"] += r["n"]
        by_status[st]["eur"] += r["total_abs_eur"] or 0
        by_department.setdefault(dept, {"matched": 0, "suggested": 0, "unmatched": 0, "manual": 0, "excluded": 0, "eur": 0.0})
        by_department[dept][st] = by_department[dept].get(st, 0) + r["n"]
        by_department[dept]["eur"] += r["total_abs_eur"] or 0
        by_holder.setdefault(holder, {"matched": 0, "suggested": 0, "unmatched": 0, "manual": 0, "excluded": 0, "eur": 0.0})
        by_holder[holder][st] = by_holder[holder].get(st, 0) + r["n"]
        by_holder[holder]["eur"] += r["total_abs_eur"] or 0
        total["n"] += r["n"]
        total["eur"] += r["total_abs_eur"] or 0

    return {
        "period": period,
        "total": {"n": total["n"], "eur": round(total["eur"], 2)},
        "by_status": {k: {"n": v["n"], "eur": round(v["eur"], 2)} for k, v in by_status.items()},
        "by_department": {k: {**v, "eur": round(v["eur"], 2)} for k, v in by_department.items()},
        "by_holder": {k: {**v, "eur": round(v["eur"], 2)} for k, v in by_holder.items()},
        "thresholds": {"auto": AUTO_MATCH_CONFIDENCE, "suggested": SUGGEST_CONFIDENCE},
    }
