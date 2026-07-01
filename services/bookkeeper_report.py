"""Bookkeeper 'Izdevumu atskaite' (expense report) — XLSX + PDF export.

Format per operator's sample (Latvia bookkeeping):

    IZDEVUMU ATSKAITE                                    <COMPANY>
    Atskaites nosaukums: <MONTH>, <YEAR>                 <COUNTRY>
    ID: <auto>
    Nozīņoja: <uploader>
    Bankas konts / IBAN: <account>
    Sākuma datums: <period start>  Beigu datums: <period end>

    ── Maksāts ar: <payment source group> ──
      Izmaksu tips: <ledger code + label>
      Nr. <group-id>
      ┌───┬──────────┬────────────┬──────────┬─────────┬──────────┬──────────┬───────┬───────┬──────┬────────┐
      │ # │ Izrakstīts │ Maksāts ar │ Komentārs │ Piegādātājs │ Dok. Nr. │ Atsauces Nr. │ Valūta │ Netto │ PVN  │ Summa  │
      ├───┼──────────┼────────────┼──────────┼─────────┼──────────┼──────────┼───────┼───────┼──────┼────────┤
      │ … per-doc rows … │
      └────────────────────────────────────────────────────────────────────────────────────────────┘
      Starpsumma                                                       EUR   NNN   NNN    NNN
    (…more Izmaksu tips groups…)

    Kopā samaksāts par uzņēmuma līdzekļiem   EUR   TOTAL  TOTAL  TOTAL
    Kopā samaksāts par uzņēmuma līdzekļiem   USD   TOTAL  TOTAL  TOTAL

Grouping order (matches the operator's PDF):
  1. Payment source (payment_method → human label)
  2. Ledger code + label (Izmaksu tips)
  3. Per-doc rows sorted by payment_executed_at ASC

Every group has a subtotal row. Report ends with a grand total per
currency across ALL groups.
"""
from __future__ import annotations

import io
import json
import logging
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

from services import db

logger = logging.getLogger(__name__)

__all__ = ["build_report_data", "generate_xlsx", "generate_pdf"]


# Human labels for our payment_method enum values, aligned to the
# operator's Latvian PDF wording.
_PAYMENT_SOURCE_LABEL_LV = {
    "bank_transfer":     "Uzņēmuma līdzekļiem",
    "wire":              "Uzņēmuma līdzekļiem",
    "sepa":              "Uzņēmuma līdzekļiem",
    "card":              "Uzņēmuma līdzekļiem",   # company card
    "company_card":      "Uzņēmuma līdzekļiem",
    "personal_card":     "Personīgiem līdzekļiem",
    "cash":              "Skaidrā naudā",
    "netting":           "Ieskaits",
    "":                  "Uzņēmuma līdzekļiem",   # unknown → default bucket
    None:                "Uzņēmuma līdzekļiem",
}

# Per-row "Maksāts ar" column (transaction-level): friendly Latvian
# label for each concrete payment channel.
_TX_METHOD_LABEL_LV = {
    "bank_transfer": "pārskaitījumu",
    "wire":          "pārskaitījumu",
    "sepa":          "pārskaitījumu",
    "card":          "uzņēmuma karti",
    "company_card":  "uzņēmuma karti",
    "personal_card": "personīgo karti",
    "cash":          "skaidrā naudā",
    "netting":       "ieskaitu",
    "":              "pārskaitījumu",
    None:            "pārskaitījumu",
}


def _parse_money_field(parsed_json: Any, field: str) -> Optional[float]:
    if isinstance(parsed_json, str):
        try:
            parsed_json = json.loads(parsed_json)
        except (ValueError, TypeError):
            return None
    if not isinstance(parsed_json, dict):
        return None
    money = parsed_json.get("money") or {}
    v = money.get(field)
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _pick_doc_number(doc: Dict[str, Any], parsed: Any) -> str:
    """Prefer the vendor's invoice number over our internal id."""
    if isinstance(parsed, str):
        try:
            parsed = json.loads(parsed)
        except (ValueError, TypeError):
            parsed = {}
    if isinstance(parsed, dict):
        inv = (parsed.get("invoice") or {}).get("number")
        if inv:
            return str(inv)
        # some parsers put it at top level
        if parsed.get("invoice_number"):
            return str(parsed["invoice_number"])
    return doc.get("original_name") or doc.get("id") or ""


def _iban_for_paying_account(paying_account_id: Optional[int]) -> str:
    if not paying_account_id:
        return ""
    conn = db.get_connection()
    try:
        row = conn.execute(
            "SELECT iban, account_name FROM paying_accounts WHERE id = ?",
            (paying_account_id,),
        ).fetchone()
    except Exception:  # noqa: BLE001 — paying_accounts may not exist in fresh DB
        return ""
    finally:
        conn.close()
    if not row:
        return ""
    return (row["iban"] or row["account_name"] or "").strip()


def _ledger_label(code: Optional[str]) -> str:
    """Look up a friendly label from ledger_schema.json for the report."""
    if not code:
        return "—"
    try:
        import config as _cfg
        with open(_cfg.LEDGER_FILE, "r", encoding="utf-8") as f:
            schema = json.load(f)
        for c in schema.get("codes", []):
            if c.get("code") == code:
                return f"{code} · {c.get('label') or c.get('group') or ''}"
    except (OSError, json.JSONDecodeError, KeyError):
        pass
    return code


def build_report_data(*,
                       period: str,
                       legal_entity: Optional[str] = None,
                       profit_center: Optional[str] = None,
                       report_id: Optional[str] = None) -> Dict[str, Any]:
    """Pull rows and shape them into the report structure.

    Returns:
      {
        "company": <legal_entity or PC>,
        "country": "Latvia" | "Estonia" | ...,
        "period_label": "Jūnijs, 2026",
        "period_start": "2026-06-01",
        "period_end":   "2026-06-30",
        "report_id":    "370799",
        "uploader":     "Artjoms Fokejevs (email)",
        "iban":         "LV..",
        "sources": [
          {
            "label": "Uzņēmuma līdzekļiem",
            "groups": [
              {
                "ledger_code": "4_1_Degviela",
                "ledger_label": "4_1_Degviela · Fuel",
                "group_id":     "202606-001",
                "rows": [
                  { "n": 1, "issue_date": "2026-05-31", "tx_method": "pārskaitījumu",
                    "comment": "", "vendor": "Circle K Latvia SIA",
                    "doc_number": "40290503", "reference": "",
                    "currency": "EUR", "net": 81.49, "vat": 17.11, "total": 98.60 },
                  …
                ],
                "subtotal": { "currency": "EUR", "net": 81.49, "vat": 17.11, "total": 98.60 },
              },
              …
            ],
            "totals_by_currency": { "EUR": {net, vat, total}, "USD": {…} },
          },
          …
        ],
        "grand_totals_by_source_and_currency": {
          "Uzņēmuma līdzekļiem": {"EUR": {…}, "USD": {…}},
          …
        },
      }
    """
    if not period or len(period) != 7:
        raise ValueError("period must be YYYY-MM")

    where = ["period = ?",
             "status IN ('paid','posted','confirmed_to_pay','approved','budget_validated','classified')"]
    params: List[Any] = [period]
    if legal_entity:
        if legal_entity == "unassigned":
            where.append("(legal_entity IS NULL OR legal_entity = '')")
        else:
            where.append("legal_entity = ?"); params.append(legal_entity)
    if profit_center:
        where.append("profit_center = ?"); params.append(profit_center)
    sql = ("SELECT * FROM documents WHERE " + " AND ".join(where)
           + " ORDER BY COALESCE(payment_executed_at, uploaded_at) ASC")
    conn = db.get_connection()
    try:
        rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    finally:
        conn.close()

    # Meta — pick the most-common uploader as "Nozīņoja"
    from collections import Counter
    up_counter = Counter((r.get("uploaded_by") or "").strip() for r in rows if r.get("uploaded_by"))
    uploader = up_counter.most_common(1)[0][0] if up_counter else ""

    # IBAN — take the most-common paying_account
    pa_counter = Counter(r.get("payment_account") for r in rows if r.get("payment_account"))
    ibans_seen = [pa for pa, _ in pa_counter.most_common()]
    iban = ""
    if ibans_seen:
        # Try to resolve first one to a real IBAN string via paying_accounts
        try:
            first = ibans_seen[0]
            pa_id = int(first) if str(first).isdigit() else None
            iban = _iban_for_paying_account(pa_id) or (str(first) if first else "")
        except (TypeError, ValueError):
            iban = str(ibans_seen[0]) if ibans_seen[0] else ""

    # Period label — Latvian month name (fallback to numeric)
    _lv_months = ["Janvāris","Februāris","Marts","Aprīlis","Maijs","Jūnijs",
                   "Jūlijs","Augusts","Septembris","Oktobris","Novembris","Decembris"]
    try:
        year, month = period.split("-")
        month_i = int(month)
        period_label = f"{_lv_months[month_i-1]}, {year}"
        pd_start = f"{year}-{month.zfill(2)}-01"
        # last day
        import calendar
        last = calendar.monthrange(int(year), month_i)[1]
        pd_end = f"{year}-{month.zfill(2)}-{last:02d}"
    except (ValueError, IndexError):
        period_label = period
        pd_start = period + "-01"
        pd_end = period + "-28"

    # Group by (payment_source_label, ledger_code)
    from collections import defaultdict
    grouped: Dict[str, Dict[str, List[Dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for r in rows:
        method = (r.get("payment_method") or "").strip().lower()
        source_lbl = _PAYMENT_SOURCE_LABEL_LV.get(method, "Uzņēmuma līdzekļiem")
        ledger = r.get("ledger_code") or "—"
        grouped[source_lbl][ledger].append(r)

    def _row_view(r: Dict[str, Any], n: int) -> Dict[str, Any]:
        parsed = r.get("parsed_json")
        net = _parse_money_field(parsed, "net_amount")
        vat = _parse_money_field(parsed, "tax_amount")
        total = r.get("amount") or r.get("amount_eur") or 0
        if net is None and vat is not None and total is not None:
            net = round(float(total) - float(vat), 2)
        if net is None:
            net = float(total or 0)
        if vat is None:
            vat = 0.0
        issue_iso = ""
        for k in ("payment_executed_at", "posted_at", "approved_at", "uploaded_at"):
            v = r.get(k)
            if v:
                issue_iso = v[:10]
                break
        method = (r.get("payment_method") or "").strip().lower()
        tx_label = _TX_METHOD_LABEL_LV.get(method, "pārskaитījumu")
        # Prefer DD/MM/YYYY per operator's PDF
        issue_disp = issue_iso
        try:
            d = datetime.strptime(issue_iso, "%Y-%m-%d")
            issue_disp = d.strftime("%d/%m/%Y")
        except (ValueError, TypeError):
            pass
        parsed_dict = parsed if isinstance(parsed, dict) else {}
        return {
            "n":           n,
            "issue_date":  issue_disp,
            "tx_method":   tx_label,
            "comment":     (r.get("payment_comment") or r.get("cost_reason") or "").strip(),
            "vendor":      (r.get("vendor") or "—").strip(),
            "doc_number":  _pick_doc_number(r, parsed),
            "reference":   (r.get("payment_reference") or "").strip(),
            "currency":    (r.get("currency_orig") or r.get("currency") or "EUR").upper(),
            "net":         round(float(net or 0), 2),
            "vat":         round(float(vat or 0), 2),
            "total":       round(float(total or 0), 2),
        }

    # Assemble
    sources_out = []
    grand_by_src_ccy: Dict[str, Dict[str, Dict[str, float]]] = defaultdict(lambda: defaultdict(lambda: {"net":0.0,"vat":0.0,"total":0.0}))
    group_seq = 0
    ym = period.replace("-", "")   # e.g. '202606'
    for source_lbl in sorted(grouped.keys()):
        groups_out = []
        src_totals_ccy: Dict[str, Dict[str, float]] = defaultdict(lambda: {"net":0.0,"vat":0.0,"total":0.0})
        for ledger in sorted(grouped[source_lbl].keys()):
            group_seq += 1
            row_views = []
            for i, r in enumerate(grouped[source_lbl][ledger], start=1):
                v = _row_view(r, i)
                row_views.append(v)
                src_totals_ccy[v["currency"]]["net"]   += v["net"]
                src_totals_ccy[v["currency"]]["vat"]   += v["vat"]
                src_totals_ccy[v["currency"]]["total"] += v["total"]
                grand_by_src_ccy[source_lbl][v["currency"]]["net"]   += v["net"]
                grand_by_src_ccy[source_lbl][v["currency"]]["vat"]   += v["vat"]
                grand_by_src_ccy[source_lbl][v["currency"]]["total"] += v["total"]
            sub_by_ccy: Dict[str, Dict[str, float]] = defaultdict(lambda: {"net":0.0,"vat":0.0,"total":0.0})
            for v in row_views:
                sub_by_ccy[v["currency"]]["net"]   += v["net"]
                sub_by_ccy[v["currency"]]["vat"]   += v["vat"]
                sub_by_ccy[v["currency"]]["total"] += v["total"]
            subs = [{
                "currency": ccy,
                "net":     round(vals["net"], 2),
                "vat":     round(vals["vat"], 2),
                "total":   round(vals["total"], 2),
            } for ccy, vals in sub_by_ccy.items()]
            groups_out.append({
                "ledger_code":  ledger,
                "ledger_label": _ledger_label(ledger),
                "group_id":     f"{ym}-{group_seq:03d}",
                "rows":         row_views,
                "subtotals":    subs,
            })
        sources_out.append({
            "label": source_lbl,
            "groups": groups_out,
            "totals_by_currency": {ccy: {k: round(v, 2) for k, v in vals.items()}
                                    for ccy, vals in src_totals_ccy.items()},
        })

    return {
        "company":       legal_entity or profit_center or "",
        "country":       "Latvia",  # TODO: derive from legal_entities.json
        "period_label":  period_label,
        "period_start":  _dt_display(pd_start),
        "period_end":    _dt_display(pd_end),
        "report_id":     report_id or ym + "-" + str(len(rows)).zfill(3),
        "uploader":      uploader,
        "iban":          iban,
        "sources":       sources_out,
        "grand_totals_by_source_and_currency": {
            src: {ccy: {k: round(v, 2) for k, v in vals.items()}
                    for ccy, vals in by_ccy.items()}
            for src, by_ccy in grand_by_src_ccy.items()
        },
        "row_count":     len(rows),
    }


def _dt_display(iso: str) -> str:
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return iso


# ────────────────────────────────────────────────────────────────────────
# XLSX generator
# ────────────────────────────────────────────────────────────────────────

def generate_xlsx(data: Dict[str, Any]) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Izdevumu atskaite"

    thin = Side(style="thin", color="D0D0D0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font  = Font(name="Calibri", size=18, bold=True, color="1E293B")
    company_font = Font(name="Calibri", size=13, bold=True, color="1E293B")
    header_font = Font(name="Calibri", size=10, bold=True, color="1E293B")
    subhead_font= Font(name="Calibri", size=11, bold=True, color="0F172A")
    small_muted = Font(name="Calibri", size=9, color="64748B")
    small_bold  = Font(name="Calibri", size=10, bold=True)
    subtotal_fill = PatternFill("solid", fgColor="FEF3C7")
    header_fill   = PatternFill("solid", fgColor="F1F5F9")

    # Header block
    ws["A1"] = "Izdevumu atskaite"
    ws["A1"].font = title_font
    ws["J1"] = data.get("company", "")
    ws["J1"].font = company_font
    ws["J1"].alignment = Alignment(horizontal="right")
    ws["J2"] = data.get("country", "")
    ws["J2"].font = small_muted
    ws["J2"].alignment = Alignment(horizontal="right")

    ws["A3"] = "Atskaites nosaukums:"; ws["A3"].font = small_bold
    ws["B3"] = data.get("period_label", "")
    ws["A4"] = "ID:";                  ws["A4"].font = small_bold
    ws["B4"] = data.get("report_id", "")
    ws["A5"] = "Nozīņoja:";            ws["A5"].font = small_bold
    ws["B5"] = data.get("uploader", "")
    ws["A6"] = "Bankas konts (IBAN):"; ws["A6"].font = small_bold
    ws["B6"] = data.get("iban", "")
    ws["A7"] = "Sākuma datums:";       ws["A7"].font = small_bold
    ws["B7"] = data.get("period_start", "")
    ws["D7"] = "Beigu datums:";        ws["D7"].font = small_bold
    ws["E7"] = data.get("period_end", "")

    row = 9
    for source in data.get("sources", []):
        ws.cell(row=row, column=1, value="Maksāts ar:").font = small_bold
        ws.cell(row=row, column=2, value=source["label"]).font = subhead_font
        row += 2

        for grp in source.get("groups", []):
            ws.cell(row=row, column=1, value="Izmaksu tips:").font = small_bold
            ws.cell(row=row, column=2, value=grp["ledger_label"]).font = subhead_font
            row += 1
            ws.cell(row=row, column=1, value="Nr.").font = small_bold
            ws.cell(row=row, column=2, value=grp["group_id"]).font = Font(name="Calibri", size=10)
            row += 1

            headers = ["#", "Izrakstīts", "Maksāts ar", "Komentārs",
                        "Piegādātājs", "Dokumenta Nr.", "Atsauces Nr.",
                        "Valūta", "Netto", "PVN", "Summa"]
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = box
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row += 1

            for rv in grp.get("rows", []):
                vals = [rv["n"], rv["issue_date"], rv["tx_method"], rv["comment"],
                         rv["vendor"], rv["doc_number"], rv["reference"],
                         rv["currency"], rv["net"], rv["vat"], rv["total"]]
                for c, v in enumerate(vals, start=1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border = box
                    if c >= 9:
                        cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                row += 1

            for sub in grp.get("subtotals", []):
                ws.cell(row=row, column=7, value="Starpsumma").font = small_bold
                ws.cell(row=row, column=8, value=sub["currency"]).font = small_bold
                for c, k in ((9, "net"), (10, "vat"), (11, "total")):
                    cell = ws.cell(row=row, column=c, value=sub[k])
                    cell.font = small_bold
                    cell.fill = subtotal_fill
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                row += 1
            row += 1
        row += 1

    # Grand totals per source per currency
    ws.cell(row=row, column=1, value="Kopā samaksāts:").font = subhead_font
    row += 1
    for src, by_ccy in (data.get("grand_totals_by_source_and_currency") or {}).items():
        for ccy, vals in by_ccy.items():
            ws.cell(row=row, column=2, value=f"par {src.lower()}").font = small_bold
            ws.cell(row=row, column=8, value=ccy).font = small_bold
            for c, k in ((9, "net"), (10, "vat"), (11, "total")):
                cell = ws.cell(row=row, column=c, value=vals[k])
                cell.font = small_bold
                cell.fill = subtotal_fill
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            row += 1

    widths = [4, 12, 14, 22, 22, 16, 16, 8, 10, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A9"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────────────────────────────────────────
# PDF generator
# ────────────────────────────────────────────────────────────────────────

def generate_pdf(data: Dict[str, Any]) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, PageBreak)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("title", parent=styles["Heading1"],
                                fontName="Helvetica-Bold", fontSize=18,
                                textColor=colors.HexColor("#1E293B"),
                                spaceAfter=4)
    company_st = ParagraphStyle("company", parent=styles["Normal"],
                                 fontName="Helvetica-Bold", fontSize=13,
                                 alignment=2, textColor=colors.HexColor("#1E293B"))
    country_st = ParagraphStyle("country", parent=styles["Normal"],
                                 fontName="Helvetica", fontSize=9,
                                 alignment=2, textColor=colors.HexColor("#64748B"),
                                 spaceAfter=8)
    meta_st = ParagraphStyle("meta", parent=styles["Normal"],
                              fontName="Helvetica", fontSize=9,
                              textColor=colors.HexColor("#0F172A"),
                              leading=13)
    subhead_st = ParagraphStyle("subhead", parent=styles["Normal"],
                                 fontName="Helvetica-Bold", fontSize=11,
                                 textColor=colors.HexColor("#0F172A"),
                                 spaceBefore=8, spaceAfter=2)
    group_meta_st = ParagraphStyle("gmeta", parent=styles["Normal"],
                                    fontName="Helvetica", fontSize=9,
                                    textColor=colors.HexColor("#64748B"),
                                    spaceAfter=4)

    story: List[Any] = []
    # Header row (title left, company right) via a 2-col table
    hdr = Table([
        [Paragraph("<b>Izdevumu atskaite</b>", title_st),
         Paragraph(data.get("company", ""), company_st)],
        ["", Paragraph(data.get("country", ""), country_st)],
    ], colWidths=[110 * mm, 65 * mm])
    hdr.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(hdr)

    meta_lines = [
        f"<b>Atskaites nosaukums:</b> {data.get('period_label','')}",
        f"<b>ID:</b> {data.get('report_id','')}",
        f"<b>Nozīņoja:</b> {data.get('uploader','')}",
        f"<b>Bankas konts</b>",
        f"&nbsp;&nbsp;<b>IBAN:</b> {data.get('iban','')}",
        f"<b>Sākuma datums:</b> {data.get('period_start','')}  &nbsp;&nbsp;"
        f"<b>Beigu datums:</b> {data.get('period_end','')}",
    ]
    for ln in meta_lines:
        story.append(Paragraph(ln, meta_st))
    story.append(Spacer(1, 8))

    col_widths = [8*mm, 20*mm, 20*mm, 28*mm, 30*mm, 22*mm, 20*mm, 12*mm, 14*mm, 12*mm, 14*mm]
    header_row = ["#", "Izrakstīts", "Maksāts ar", "Komentārs",
                   "Piegādātājs", "Dokumenta Nr.", "Atsauces Nr.",
                   "Valūta", "Netto", "PVN", "Summa"]

    for source in data.get("sources", []):
        story.append(Paragraph(f"<b>Maksāts ar:</b> {source['label']}", subhead_st))

        for grp in source.get("groups", []):
            story.append(Paragraph(f"<b>Izmaksu tips:</b> {grp['ledger_label']}", subhead_st))
            story.append(Paragraph(f"Nr. {grp['group_id']}", group_meta_st))

            table_data: List[List[Any]] = [header_row]
            for rv in grp.get("rows", []):
                table_data.append([
                    rv["n"],
                    rv["issue_date"],
                    rv["tx_method"],
                    rv["comment"] or "",
                    rv["vendor"],
                    rv["doc_number"],
                    rv["reference"] or "",
                    rv["currency"],
                    f"{rv['net']:.2f}",
                    f"{rv['vat']:.2f}",
                    f"{rv['total']:.2f}",
                ])
            # Subtotal rows per currency
            for sub in grp.get("subtotals", []):
                table_data.append([
                    "", "", "", "", "", "", "Starpsumma",
                    sub["currency"],
                    f"{sub['net']:.2f}",
                    f"{sub['vat']:.2f}",
                    f"{sub['total']:.2f}",
                ])

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#F1F5F9")),
                ("TEXTCOLOR",    (0,0), (-1,0), colors.HexColor("#1E293B")),
                ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 8),
                ("VALIGN",       (0,0), (-1,-1), "TOP"),
                ("ALIGN",        (0,0), (0,-1),  "CENTER"),
                ("ALIGN",        (7,0), (-1,-1), "RIGHT"),
                ("GRID",         (0,0), (-1,-1), 0.25, colors.HexColor("#D0D0D0")),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
                ("TOPPADDING",   (0,0), (-1,-1), 3),
                # Subtotal rows: amber background
                *[("BACKGROUND", (0, -1-i), (-1, -1-i), colors.HexColor("#FEF3C7"))
                   for i in range(len(grp.get("subtotals", [])))],
                *[("FONTNAME",   (0, -1-i), (-1, -1-i), "Helvetica-Bold")
                   for i in range(len(grp.get("subtotals", [])))],
            ]))
            story.append(tbl)
            story.append(Spacer(1, 8))

    # Grand totals block
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>Kopā samaksāts:</b>", subhead_st))
    for src, by_ccy in (data.get("grand_totals_by_source_and_currency") or {}).items():
        for ccy, vals in by_ccy.items():
            grand = [[
                f"Kopā samaksāts par {src.lower()}",
                ccy,
                f"{vals['net']:.2f}",
                f"{vals['vat']:.2f}",
                f"{vals['total']:.2f}",
            ]]
            gt = Table(grand, colWidths=[90*mm, 15*mm, 22*mm, 22*mm, 22*mm])
            gt.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), colors.HexColor("#FEF3C7")),
                ("FONTNAME",     (0,0), (-1,-1), "Helvetica-Bold"),
                ("FONTSIZE",     (0,0), (-1,-1), 10),
                ("ALIGN",        (1,0), (-1,-1), "RIGHT"),
                ("BOX",          (0,0), (-1,-1), 0.5, colors.HexColor("#D97706")),
                ("BOTTOMPADDING",(0,0), (-1,-1), 6),
                ("TOPPADDING",   (0,0), (-1,-1), 5),
            ]))
            story.append(gt)
            story.append(Spacer(1, 4))

    doc.build(story)
    return buf.getvalue()
