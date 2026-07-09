"""Document parsing service using Claude Vision, pypdf, and tabular parsers."""
from __future__ import annotations

import base64
import csv
import io
import json
import logging
import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional

import config
from services import fx  # Phase 2.1 — ECB EUR conversion

logger = logging.getLogger(__name__)

__all__ = ["parse_document"]

# Currencies for which an invoice MUST be converted to EUR before posting.
# (Anything that's not EUR-zone — Lilya's USD case, Denis's Yerevan AMD case.)
_EU_EURO_ZONE = {
    "EUR", "€",
}
# Non-EU country hints — when these appear in vendor address, force currency check
_NON_EU_COUNTRY_HINTS = (
    "armenia", "yerevan", "ереван", "армения",  # AMD
    "georgia", "tbilisi", "тбилиси", "грузия",   # GEL
    "uk", "united kingdom", "london", "england", # GBP
    "usa", "us", "united states", "new york",    # USD
    "switzerland", "geneva", "zurich",            # CHF
    "russia", "moscow", "россия", "москва",       # RUB
    "ukraine", "kyiv", "kiev", "украина",        # UAH
    "turkey", "istanbul",                         # TRY
    "uae", "dubai", "abu dhabi",                  # AED
)


def _enrich_with_fx(parsed: Any) -> Any:
    """Add EUR equivalent to parsed['money'] via ECB rates.

    Works for both single-doc dict and multi-doc list. Idempotent — re-runs
    safely (skips when amount_eur already present and currency hasn't changed).
    """
    if isinstance(parsed, list):
        return [_enrich_with_fx(d) for d in parsed]
    if not isinstance(parsed, dict):
        return parsed
    money = parsed.get("money") or {}
    amount = money.get("total_amount")
    if amount is None:
        return parsed
    currency = (money.get("currency") or "EUR").upper().strip()
    # Sanity check — if vendor is in non-EU country but currency is EUR,
    # raise a warning (testers Denis/Lilya complained this is exact pattern)
    addr = ""
    vendor = parsed.get("vendor") or {}
    if isinstance(vendor, dict):
        addr = (vendor.get("address") or "").lower()
    if currency == "EUR" and any(h in addr for h in _NON_EU_COUNTRY_HINTS):
        warnings = parsed.setdefault("warnings", [])
        warnings.append(
            "Currency=EUR but vendor address suggests non-EU country — "
            "please verify currency on document (auto-detect may have failed)"
        )
    doc_date = (parsed.get("dates") or {}).get("document_date")
    fx_data = fx.convert_to_eur(amount, currency, doc_date)
    money.update(fx_data)
    parsed["money"] = money
    return parsed


def parse_document(file_path: str, file_type: str) -> Dict[str, Any]:
    """Parse a financial document and extract structured data.

    Args:
        file_path: Absolute path to the uploaded file.
        file_type: File extension (pdf, jpg, png, csv, xlsx).

    Returns:
        Extracted data dictionary with keys: document_type, vendor, dates,
        money (with EUR equivalent + FX metadata), line_items, ledger_codes, warnings.
    """
    logger.info("Parsing %s (type=%s)", file_path, file_type)

    if file_type in ("csv",):
        result = _parse_csv(file_path)
    elif file_type in ("xlsx",):
        result = _parse_xlsx(file_path)
    elif file_type in ("pdf",):
        result = _parse_pdf(file_path)
    elif file_type in ("jpg", "jpeg", "png"):
        result = _parse_image(file_path)
    else:
        return _mock_response(file_path, "Unsupported file type: %s" % file_type)

    # Phase 2.1 — always enrich with FX (ECB EUR equivalent)
    result = _enrich_with_fx(result)
    # Multi-source registry: VIES for EU, OpenCorporates fallback for non-EU
    try:
        from services.company_registry import registry_enrich_vendor
        result = registry_enrich_vendor(result)
    except Exception as exc:
        logger.debug("registry enrich skipped: %s", exc)
    return result


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _parse_pdf(file_path: str) -> Dict[str, Any]:
    """Parse a PDF -- try text extraction first, fall back to Vision."""
    text = _extract_pdf_text(file_path)
    if text and len(text.strip()) > 50:
        result = _parse_text_with_llm(text, os.path.basename(file_path))
        # Enrich vendor with regex-extracted details (VAT, Reg.Nr, IBAN, address)
        # This works even if LLM failed and returned mock data
        result = _enrich_with_regex_details(result, text)
        return result
    # Fall back to Vision (treat each page as an image)
    return _parse_image(file_path)


# ---------------------------------------------------------------------------
# Regex-based VAT / Reg.Nr / IBAN / Address extraction
# Works without LLM, enriches LLM output, salvages failed LLM calls.
# ---------------------------------------------------------------------------

# EU VAT formats per https://en.wikipedia.org/wiki/VAT_identification_number
_VAT_PATTERNS = [
    (r"\b(LV)\s*[-]?\s*(\d{11})\b", "LV"),
    (r"\b(EE)\s*[-]?\s*(\d{9})\b", "EE"),
    (r"\b(LT)\s*[-]?\s*(\d{9}|\d{12})\b", "LT"),
    (r"\b(PL)\s*[-]?\s*(\d{10})\b", "PL"),
    (r"\b(DE)\s*[-]?\s*(\d{9})\b", "DE"),
    (r"\b(FR)\s*[-]?\s*([A-HJ-NP-Z0-9]{2}\d{9})\b", "FR"),
    (r"\b(GB)\s*[-]?\s*(\d{9}|\d{12})\b", "GB"),
    (r"\b(IE)\s*[-]?\s*(\d{7}[A-Z]{1,2}|\d[A-Z\d]\d{5}[A-Z])\b", "IE"),
    (r"\b(NL)\s*[-]?\s*(\d{9}B\d{2})\b", "NL"),
    (r"\b(BE)\s*[-]?\s*(0?\d{9})\b", "BE"),
    (r"\b(IT)\s*[-]?\s*(\d{11})\b", "IT"),
    (r"\b(ES)\s*[-]?\s*([A-Z]\d{7}[A-Z0-9])\b", "ES"),
    (r"\b(SE)\s*[-]?\s*(\d{12})\b", "SE"),
    (r"\b(FI)\s*[-]?\s*(\d{8})\b", "FI"),
    (r"\b(DK)\s*[-]?\s*(\d{8})\b", "DK"),
    (r"\b(AT)\s*[-]?\s*(U\d{8})\b", "AT"),
    (r"\b(CZ)\s*[-]?\s*(\d{8,10})\b", "CZ"),
    (r"\b(SK)\s*[-]?\s*(\d{10})\b", "SK"),
    (r"\b(HU)\s*[-]?\s*(\d{8})\b", "HU"),
    (r"\b(BG)\s*[-]?\s*(\d{9,10})\b", "BG"),
    (r"\b(RO)\s*[-]?\s*(\d{2,10})\b", "RO"),
    (r"\b(HR)\s*[-]?\s*(\d{11})\b", "HR"),
    (r"\b(SI)\s*[-]?\s*(\d{8})\b", "SI"),
    (r"\b(LU)\s*[-]?\s*(\d{8})\b", "LU"),
    (r"\b(MT)\s*[-]?\s*(\d{8})\b", "MT"),
    (r"\b(CY)\s*[-]?\s*(\d{8}[A-Z])\b", "CY"),
    (r"\b(PT)\s*[-]?\s*(\d{9})\b", "PT"),
    (r"\b(GR|EL)\s*[-]?\s*(\d{9})\b", "EL"),
]

# Latvian "Reg. Nr." patterns (without LV prefix) -- often appears separately
_LV_REG_PATTERNS = [
    r"(?:PVN\s*)?[Rr]eg\.?\s*[Nn]r\.?[:\s]*(\d{11})",
    r"[Nn]odoklu\s+maks\.?\s*reg\.?\s*nr\.?[:\s]*(\d{11})",
    r"[Vv][AaĀā][Tt]\s*(?:[Nn]o\.?|[Nn]umber)?[:\s]*(LV\d{11})",
    r"[Tt]ax\s*ID[:\s]*(LV\d{11})",
]

_IBAN_PATTERN = r"\b([A-Z]{2}\d{2}\s?(?:[A-Z0-9]{4}\s?){3,7}[A-Z0-9]{1,4})\b"
_LV_POSTAL = r"\bLV\s*[-]?\s*(\d{4})\b"
_ADDRESS_HINTS = (
    r"(?:iela|street|str\.?|prospekts|aleja|laukums|bulvaris|"
    r"strasse|straße|ulica|avenue|ave\.?|road|rd\.?)"
)


def _normalise_digit_runs(text: str) -> str:
    """Collapse whitespace / thin-spaces inside digit groups so a VAT
    number like "LV 4540 3022 922" or "LV45 403 022 922" survives OCR
    that put stray spaces between the digits.

    2026-07-02 op-feedback — parser missed LV45403022922 on the Sipalto
    invoice because the PDF text layer split the VAT with spaces
    ("LV  45 40 30 22 92 2"). We normalise a copy of the text before
    running the strict VAT regexes; original text is kept for
    downstream extraction (address hints, IBAN with spaces).
    """
    # Collapse runs of whitespace between digits: "45 40 30 22 92 2" → "454030229 2"
    # then repeat until stable (handles "1 2 3 4" → "1234").
    prev = None
    cur = text
    while prev != cur:
        prev = cur
        cur = re.sub(r"(\d)[ \t  ]+(\d)", r"\1\2", cur)
    # Also collapse whitespace between country prefix and digits
    cur = re.sub(r"\b([A-Z]{2})[ \t  ]+(\d)", r"\1\2", cur)
    return cur


def _extract_vat_details(text: str) -> Dict[str, Any]:
    """Extract VAT, registration number, IBAN, postal code, address hints from raw text.

    Returns a dict with optional keys: vat_number, registration_number, iban,
    postal_code, address_hint. Missing fields are omitted.
    """
    out: Dict[str, Any] = {}

    # Run VAT-format regexes against a whitespace-normalised copy so
    # broken OCR spacing does not sink an otherwise obvious VAT.
    search_text = _normalise_digit_runs(text)

    # 1. EU VAT
    for pattern, country in _VAT_PATTERNS:
        m = re.search(pattern, search_text, re.IGNORECASE)
        if m:
            out["vat_number"] = (country + m.group(2)).upper().replace(" ", "").replace("-", "")
            out["vat_country"] = country
            break

    # 2. Latvian Reg.Nr. specifically — also on normalised text
    for pattern in _LV_REG_PATTERNS:
        m = re.search(pattern, search_text)
        if m:
            digits = m.group(1)
            if digits.startswith("LV"):
                out.setdefault("vat_number", digits)
                out["registration_number"] = digits[2:]
            else:
                out["registration_number"] = digits
                out.setdefault("vat_number", "LV" + digits)
            break

    # 3. IBAN
    m = re.search(_IBAN_PATTERN, text)
    if m:
        out["iban"] = re.sub(r"\s+", "", m.group(1))

    # 4. Latvian postal code
    m = re.search(_LV_POSTAL, text)
    if m:
        out["postal_code"] = "LV-" + m.group(1)

    # 5. Address hint (line containing street keyword)
    for line in text.split("\n"):
        if re.search(_ADDRESS_HINTS, line, re.IGNORECASE) and 5 < len(line.strip()) < 200:
            out["address_hint"] = line.strip()
            break

    return out


def _enrich_with_regex_details(result: Any, raw_text: str) -> Any:
    """Merge regex-extracted vendor details into the parsed result, then call VIES.

    Pipeline:
      1. Regex pulls VAT/IBAN/postal/address from raw text.
      2. Merge into vendor dict (LLM values win where present).
      3. If a VAT was found, hit VIES for the official name + legal address.
    """
    details = _extract_vat_details(raw_text)

    def _merge_into_doc(doc: Dict[str, Any]) -> Dict[str, Any]:
        vendor = doc.get("vendor")
        if isinstance(vendor, str):
            vendor = {"name": vendor}
        elif not isinstance(vendor, dict):
            vendor = {}
        for key, value in details.items():
            if not vendor.get(key):
                vendor[key] = value
        # VIES enrichment runs even without regex hits, in case LLM extracted VAT
        try:
            from services.vies import vies_enrich_vendor
            vendor = vies_enrich_vendor(vendor)
        except Exception as exc:
            logger.warning("VIES enrichment failed: %s", exc)
        doc["vendor"] = vendor
        return doc

    if isinstance(result, list):
        return [_merge_into_doc(d) if isinstance(d, dict) else d for d in result]
    if isinstance(result, dict):
        return _merge_into_doc(result)
    return result


def _extract_pdf_text(file_path: str) -> str:
    """Extract text from a PDF using pypdf."""
    try:
        from pypdf import PdfReader

        reader = PdfReader(file_path)
        pages: List[str] = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
        return "\n\n".join(pages)
    except Exception as exc:
        logger.warning("pypdf extraction failed: %s", exc)
        return ""


# ---------------------------------------------------------------------------
# Image (Claude Vision)
# ---------------------------------------------------------------------------

def _parse_image(file_path: str) -> Dict[str, Any]:
    """Parse an image or PDF via Claude Vision API."""
    if not config.ANTHROPIC_API_KEY:
        logger.warning("No ANTHROPIC_API_KEY set -- returning mock data")
        return _mock_response(file_path, "No API key configured (mock mode)")

    try:
        import anthropic

        with open(file_path, "rb") as f:
            raw = f.read()

        ext = os.path.splitext(file_path)[1].lower().lstrip(".")
        media_map = {
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
            "pdf": "application/pdf",
        }
        media_type = media_map.get(ext, "application/octet-stream")
        b64 = base64.b64encode(raw).decode("utf-8")

        client = anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY, timeout=60.0, max_retries=1)

        content_block: List[Dict[str, Any]]
        if ext == "pdf":
            content_block = [
                {
                    "type": "document",
                    "source": {
                        "type": "base64",
                        "media_type": "application/pdf",
                        "data": b64,
                    },
                },
                {"type": "text", "text": _extraction_prompt()},
            ]
        else:
            content_block = [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": media_type,
                        "data": b64,
                    },
                },
                {"type": "text", "text": _extraction_prompt()},
            ]

        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            messages=[{"role": "user", "content": content_block}],
        )

        text = response.content[0].text
        result = _parse_llm_json(text, file_path)
        # Run VIES on whatever VAT the LLM extracted from the image
        result = _enrich_with_regex_details(result, "")
        return result

    except Exception as exc:
        logger.exception("Vision parsing failed: %s", exc)
        return _mock_response(file_path, "Vision parsing error: %s" % str(exc))


# ---------------------------------------------------------------------------
# Text-based LLM extraction
# ---------------------------------------------------------------------------

def _parse_text_with_llm(text: str, filename: str) -> Dict[str, Any]:
    """Send extracted text to Claude for structured extraction."""
    if not config.ANTHROPIC_API_KEY:
        logger.warning("No ANTHROPIC_API_KEY set -- returning mock data")
        return _mock_response(filename, "No API key configured (mock mode)")

    try:
        import anthropic

        client = anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY, timeout=60.0, max_retries=1)
        prompt = _extraction_prompt() + "\n\nDocument text:\n" + text[:8000]

        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}],
        )

        return _parse_llm_json(response.content[0].text, filename)

    except Exception as exc:
        logger.exception("Text LLM parsing failed: %s", exc)
        return _mock_response(filename, "LLM parsing error: %s" % str(exc))


# ---------------------------------------------------------------------------
# CSV / XLSX
# ---------------------------------------------------------------------------

def _parse_csv(file_path: str) -> Dict[str, Any]:
    """Parse a CSV bank or invoice export."""
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        line_items: List[Dict[str, Any]] = []
        total = 0.0
        vendor = ""
        for row in rows[:50]:
            amount = _extract_amount_from_row(row)
            total += amount
            desc = " ".join(str(v) for v in row.values() if v)
            if not vendor:
                vendor = _guess_vendor_from_text(desc)
            line_items.append({
                "description": desc[:200],
                "amount": amount,
                "raw": {k: v for k, v in row.items()},
            })

        return {
            "document_type": "bank_statement" if len(rows) > 3 else "invoice",
            "vendor": vendor,
            "dates": {"document_date": datetime.utcnow().strftime("%Y-%m-%d")},
            "money": {
                "total_amount": round(total, 2),
                "currency": "EUR",
                "line_count": len(line_items),
            },
            "line_items": line_items,
            "ledger_codes": [],
            "warnings": [],
        }
    except Exception as exc:
        logger.exception("CSV parsing failed: %s", exc)
        return _mock_response(file_path, "CSV parse error: %s" % str(exc))


def _parse_xlsx(file_path: str) -> Dict[str, Any]:
    """Parse an XLSX file."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return _mock_response(file_path, "Empty workbook")

        headers: List[str] = []
        line_items: List[Dict[str, Any]] = []
        total = 0.0
        vendor = ""

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
                continue
            if i > 50:
                break
            row_dict = {
                headers[j]: (cell if cell is not None else "")
                for j, cell in enumerate(row)
                if j < len(headers)
            }
            amount = _extract_amount_from_row(row_dict)
            total += amount
            desc = " ".join(str(v) for v in row_dict.values() if v)
            if not vendor:
                vendor = _guess_vendor_from_text(desc)
            line_items.append({
                "description": desc[:200],
                "amount": amount,
                "raw": row_dict,
            })

        wb.close()

        return {
            "document_type": "spreadsheet",
            "vendor": vendor,
            "dates": {"document_date": datetime.utcnow().strftime("%Y-%m-%d")},
            "money": {
                "total_amount": round(total, 2),
                "currency": "EUR",
                "line_count": len(line_items),
            },
            "line_items": line_items,
            "ledger_codes": [],
            "warnings": [],
        }
    except Exception as exc:
        logger.exception("XLSX parsing failed: %s", exc)
        return _mock_response(file_path, "XLSX parse error: %s" % str(exc))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extraction_prompt() -> str:
    """Return the system prompt for financial document extraction."""
    return (
        "You are a financial document parser for Amitours Holding (Latvia). "
        "The file may contain MULTIPLE receipts, invoices or pages — treat each invoice/receipt "
        "as a SEPARATE document. Multi-page PDFs with different invoice numbers on each page = multiple documents.\n\n"
        "CRITICAL: Return a JSON ARRAY. One object per distinct invoice/receipt. "
        "Even for a single document, return an array with one element.\n\n"
        "Return ONLY valid JSON (no markdown fences). Format:\n\n"
        "[\n"
        "  {\n"
        '    "document_type": "invoice" | "receipt" | "bank_tx" | "credit_note" | "terminal_receipt" | "other",\n'
        '    "vendor": {\n'
        '      "name": "SIA Example Company",\n'
        '      "vat_number": "LV40203241255",\n'
        '      "address": "Street 1, Riga, Latvia",\n'
        '      "country": "LV"\n'
        "    },\n"
        '    "dates": {"document_date": "YYYY-MM-DD", "due_date": null},\n'
        '    "money": {\n'
        '      "subtotal": 1400.00,\n'
        '      "discount": 0.00,\n'
        '      "credits": 10.00,\n'
        '      "tax_amount": 0.00,\n'
        '      "net_amount": 1400.00,\n'
        '      "total_amount": 1390.00,\n'
        '      "currency": "EUR"\n'
        "    },\n"
        '    "line_items": [\n'
        "      {\n"
        '        "description_original": "Imeretiesu hacapuri",\n'
        '        "description_en": "Imereti khachapuri (Georgian cheese bread)",\n'
        '        "description": "Imereti khachapuri (Georgian cheese bread)",\n'
        '        "amount": 9.50,\n'
        '        "quantity": 1,\n'
        '        "language": "lv",\n'
        '        "category_hint": "food" | "travel" | "marketing" | "office" | "subscription" | "consulting" | "other"\n'
        "      }\n"
        "    ],\n"
        '    "payment_method": "card" | "bank" | "cash" | "unknown",\n'
        '    "is_terminal_receipt": false,\n'
        '    "linked_vendor": null,\n'
        '    "our_entity": "ALPS2ALPS_OU" | "AMITOURS_LONDON" | "LORA_TRANS" | "DMS" | "AMITOURS_GROUP_SA" | "ALEXCURSION" | "SINKOPA" | "AMITOURS_HOLDING" | "AMITOURS_CORP" | null,\n'
        '    "ledger_codes": [],\n'
        '    "warnings": []\n'
        "  }\n"
        "]\n\n"
        "═══ OUR ENTITY (Top-10 P2.1, Rita 2026-06-09) ═══\n"
        "- Look at the BILL TO / BUYER section of the invoice to identify which Amitours entity is being billed.\n"
        "- Map the recipient name → one of the codes above. Heuristics:\n"
        "  · 'Alps2Alps OÜ' / 'Alps2Alps OU' / 'A2A OÜ' → ALPS2ALPS_OU\n"
        "  · 'Amitours London' / 'Amitours London Ltd' → AMITOURS_LONDON\n"
        "  · 'Lora Trans' / 'Lora Trans Kft' → LORA_TRANS\n"
        "  · 'DMS' / 'DMS SIA' → DMS\n"
        "  · 'Amitours Group SA' → AMITOURS_GROUP_SA\n"
        "  · 'Alexcursion' → ALEXCURSION\n"
        "  · 'Sinkopa' → SINKOPA\n"
        "  · 'Amitours Holding' / 'Amitours Holding OÜ' → AMITOURS_HOLDING\n"
        "  · 'Amitours Corp' / 'Amitours Inc' → AMITOURS_CORP\n"
        "- If the bill-to section is missing or doesn't match any of these → null. Do NOT guess.\n\n"
        "IMPORTANT RULES:\n\n"
        "═══ MONEY BREAKDOWN (Phase 2.5 — Denis Bazoom case) ═══\n"
        "- subtotal = sum of line items BEFORE tax, BEFORE discounts/credits\n"
        "- discount = explicit discount amount (negative reduction on subtotal)\n"
        "- credits = 'credits used', 'voucher', 'loyalty redemption' (separate from discount)\n"
        "- tax_amount = VAT amount\n"
        "- net_amount = subtotal − discount − credits  (amount BEFORE tax)\n"
        "- total_amount = FINAL amount the buyer pays = net_amount + tax_amount\n"
        "- Sanity check yourself: total_amount must equal (subtotal − discount − credits + tax_amount)\n"
        "- If invoice shows 'Subtotal: 1400, Credits: −10, Total: 1390' → subtotal=1400, credits=10, total=1390 (NOT net=1400!)\n\n"
        "═══ CURRENCY DETECTION (Phase 2.1 — Lilya/Denis non-EU case) ═══\n"
        "- Use ISO-4217 codes: EUR, USD, GBP, CHF, AMD (Armenian Dram), GEL (Georgian Lari), RUB, UAH, TRY, AED\n"
        "- LOOK at currency symbols on invoice: € EUR, $ USD, £ GBP, ֏ AMD, ₾ GEL, ₽ RUB\n"
        "- Look at vendor country/address: Yerevan/Armenia → likely AMD; London → GBP; New York → USD\n"
        "- DO NOT default to EUR. If currency is unclear from non-Latvian vendor — leave currency field as best guess but add a warning 'currency_uncertain'\n"
        "- For Armenian invoice (Yerevan address) with amount 22500: currency is almost certainly AMD, NOT EUR\n"
        "- For non-EUR amounts, the EUR equivalent will be computed downstream via ECB rates\n\n"
        "═══ PAYMENT METHOD (Phase 2.4 — Denis card case) ═══\n"
        "- 'card' if you see: 'paid by card', 'invoice has been paid by card', 'Visa', 'Mastercard',\n"
        "   'card ending in', 'card xxxx 1234', 'bankas karte', 'maksāts ar karti', 'POS-terminālis'\n"
        "- 'bank' if you see: 'bank transfer', 'wire', 'IBAN', 'pārskaitījums', 'pārskaitīt',\n"
        "   'pay to account', 'SEPA', 'beneficiary' — but ONLY when no card indicator is present\n"
        "- 'cash' if you see: 'cash', 'skaidra nauda', 'paid in cash'\n"
        "- 'unknown' if not stated\n"
        "- If both card AND IBAN appear, prefer 'card' (IBAN is just metadata) UNLESS document says 'pending payment via bank transfer'\n\n"
        "═══ MULTI-PAGE / MULTI-INVOICE (Phase 2.3 — Denis case) ═══\n"
        "- If PDF has multiple pages and each has a different invoice number / date / vendor → return MULTIPLE objects\n"
        "- If pages are continuation of single invoice (same number, line-items spilling) → ONE object with merged line_items\n"
        "- Hotel bookings with multiple nights ARE one invoice (not per-night)\n\n"
        "═══ LINE-ITEM CATEGORISATION (Phase 2.5 + per-line classification) ═══\n"
        "- For EACH line item, suggest category_hint to help downstream ledger routing\n"
        "- 'travel' = hotel, taxi, flight, fuel, parking — even if invoice is mixed\n"
        "- 'marketing' = ads spend, content production, design services, PR\n"
        "- 'consulting' = legal, accounting, advisory fees\n"
        "- 'office' = supplies, rent, utilities, software\n"
        "- 'subscription' = SaaS, hosting, recurring tooling\n"
        "- 'food' = meals, business dinners (often need attendee list)\n"
        "- This hint lets the classifier split one invoice across multiple ledger codes\n\n"
        "═══ VAT NUMBER EXTRACTION ═══\n"
        "- ALWAYS look for the vendor's VAT/tax registration number\n"
        "- Latvian patterns: 'PVN Reg. Nr.', 'Reg. Nr.', 'PVN reg. nr.', 'PVN Nr.', 'Nodoklu maks. reg. nr.'\n"
        "- Format is usually 'LV' followed by 11 digits, e.g. LV40203241255\n"
        "- Also look for generic 'VAT', 'Tax ID', 'TIN' patterns\n"
        "- If registration number found without 'LV' prefix but is 11 digits, add 'LV' prefix\n\n"
        "═══ LANGUAGE DETECTION & TRANSLATION ═══\n"
        "- Detect language of each line item description\n"
        "- Translate ALL line items to English in 'description_en'\n"
        "- Keep the original text in 'description_original'\n"
        "- Set 'description' to the English translation\n"
        "- Set 'language' to ISO 639-1 code (lv, ru, ka, en, hy[Armenian], etc.)\n\n"
        "═══ TERMINAL RECEIPT LINKING ═══\n"
        "- Card terminal receipts (kvitanciis/kvits) are separate from purchase receipts\n"
        "- Set 'is_terminal_receipt': true, 'document_type': 'terminal_receipt'\n"
        "- If terminal receipt appears alongside purchase receipt for SAME amount:\n"
        '  - On terminal: set "linked_vendor": {"name": "SIA Snabb", "vat_number": "LV..."}\n'
        '  - Add warning: "duplicate_payment_confirmation"\n\n'
        "═══ KNOWN-VENDOR EXTRACTION HINTS (Phase 5b — Lilya feedback) ═══\n"
        "These vendors had repeated parsing failures. Apply the rules below:\n\n"
        "• LinkedIn (Premium / Ads / Sales Navigator):\n"
        "  - Vendor name on invoice header: 'LinkedIn Ireland Unlimited Company'\n"
        "  - VAT number bottom-right block: starts with 'IE' (e.g. IE9740425P)\n"
        "  - Currency varies: EUR for EU billing, USD if US billing — check\n"
        "    'Bill to:' country, NOT the LinkedIn-Ireland header\n"
        "  - Address: 'Wilton Place, Dublin 2, Ireland' (vendor side)\n"
        "  - Total appears as 'Total due', 'Amount due', sometimes 'Charge total'\n"
        "  - Date: 'Invoice date' field, format MM/DD/YYYY (US) or DD/MM/YYYY\n"
        "  - Payment method: usually 'card' (LinkedIn auto-charges)\n\n"
        "• Fireflies.ai (meeting transcription SaaS):\n"
        "  - Vendor name: 'Fireflies AI, Inc.' (US Delaware C-corp)\n"
        "  - NO VAT number — use 'EIN' or 'Tax ID' field if present\n"
        "  - Currency: USD (default), occasionally EUR via Stripe Tax\n"
        "  - Address: '548 Market St, San Francisco, CA' (or similar US)\n"
        "  - 'Invoice number': INV-XXXXXX format\n"
        "  - Line item usually 'Pro plan' / 'Business plan' / 'Enterprise'\n"
        "  - This is a subscription → category_hint: 'subscription'\n\n"
        "• Hetzner (DE hosting):\n"
        "  - Vendor: 'Hetzner Online GmbH', VAT: DE812871812\n"
        "  - Always EUR, German invoice\n"
        "  - 'Bruttobetrag' = total, 'MwSt' = VAT 19%\n"
        "  - category_hint: 'subscription' (servers)\n\n"
        "• Google Ireland (Ads / Cloud / Workspace):\n"
        "  - Vendor: 'Google Ireland Limited', VAT: IE6388047V\n"
        "  - EUR for EU customers\n"
        "  - For Google Ads: category_hint 'marketing'\n"
        "  - For Google Cloud / Workspace: category_hint 'subscription'\n\n"
        "• Stripe (payouts / fees):\n"
        "  - Vendor: 'Stripe Payments Europe, Ltd.' or 'Stripe, Inc.'\n"
        "  - Often arrives as a CSV transactions report, not a classic invoice\n"
        "  - For fee statements: category_hint 'subscription' (payment processing)\n"
        "  - For payouts in: this is REVENUE, not an expense\n\n"
        "• Revolut Business (multi-currency cards):\n"
        "  - Vendor: 'Revolut Payments UAB' (EU) / 'Revolut Ltd' (UK)\n"
        "  - Statement format, not invoice — extract each line as separate doc\n"
        "  - 'Card transaction' rows → look up the real vendor in description\n\n"
        "• Meta / Facebook Ads:\n"
        "  - Vendor: 'Meta Platforms Ireland Limited', VAT: IE9692928F\n"
        "  - Always check Bill-to address for the buyer's entity\n"
        "  - category_hint: 'marketing'\n\n"
        "• Apple Services (Apple Developer / iCloud Business):\n"
        "  - Vendor: 'Apple Distribution International Ltd', VAT: IE9700053D\n"
        "  - Always EUR for EU\n\n"
        "═══ OTHER RULES ═══\n"
        "- Latvian receipts: 'PVN' (VAT), 'Kopā EUR' (total), 'Bez PVN' (net), 'Summa PVN' (tax)\n"
        "- Extract actual company name from 'SIA ...' or 'AS ...' headers\n"
        "- Amounts should be positive numbers\n"
        "- Date format in Latvia: DD.MM.YYYY or YYYY-MM-DD"
    )


def _parse_llm_json(text: str, filename: str) -> Any:
    """Extract JSON from LLM response text. May return list (multi-receipt) or dict."""
    cleaned = re.sub(r"```json\s*", "", text)
    cleaned = re.sub(r"```\s*$", "", cleaned.strip())
    try:
        result = json.loads(cleaned)
        # If it's a list of documents, return as-is (but never empty)
        if isinstance(result, list):
            if not result:
                logger.warning("LLM returned empty list for %s -- treating as unparseable", filename)
                return _mock_response(filename, "LLM returned empty list (likely blank/scanned/unreadable document)")
            logger.info("Multi-document parse: %d documents found in %s", len(result), filename)
            return result
        return result
    except json.JSONDecodeError:
        # Try to find JSON array first, then object
        match_arr = re.search(r"\[[\s\S]*\]", cleaned)
        if match_arr:
            try:
                result = json.loads(match_arr.group())
                if isinstance(result, list) and result:
                    return result
            except json.JSONDecodeError:
                pass
        match_obj = re.search(r"\{[\s\S]*\}", cleaned)
        if match_obj:
            try:
                return json.loads(match_obj.group())
            except json.JSONDecodeError:
                pass
        logger.warning("Could not parse LLM JSON for %s", filename)
        return _mock_response(filename, "Failed to parse LLM response")


def _extract_amount_from_row(row: Dict[str, Any]) -> float:
    """Try to find an amount value in a row dictionary."""
    amount_keys = ["amount", "total", "sum", "debit", "credit", "value", "price"]
    for key in row:
        if any(ak in str(key).lower() for ak in amount_keys):
            try:
                val = str(row[key]).replace(",", ".").replace(" ", "")
                return abs(float(val))
            except (ValueError, TypeError):
                continue
    return 0.0


def _guess_vendor_from_text(text: str) -> str:
    """Try to extract a vendor name from free text."""
    known = [
        "Google", "Facebook", "LinkedIn", "AWS", "Hetzner",
        "Stripe", "Revolut", "Sebra Auto", "PeopleForce",
    ]
    for v in known:
        if v.lower() in text.lower():
            return v
    return ""


def _mock_response(filename: str, warning: str) -> Dict[str, Any]:
    """Return an *empty* fallback when parsing fails.

    No fake numbers -- the UI must show this as 'needs manual input', not as
    real parsed data. Categorises the failure so the frontend can render a
    specific banner.
    """
    category = "parser_unknown"
    user_msg = warning
    low = (warning or "").lower()
    if "credit balance" in low or "billing" in low or "insufficient credit" in low:
        category = "llm_no_credit"
        user_msg = (
            "Anthropic API credits exhausted. The AI parser is offline; "
            "regex + VIES still try to extract VAT/address. "
            "Please fill missing fields manually until billing is topped up."
        )
    elif "no api key" in low:
        category = "llm_no_key"
        user_msg = "ANTHROPIC_API_KEY not configured on the server."
    elif "rate" in low and "limit" in low:
        category = "llm_rate_limit"
        user_msg = "AI parser rate-limited. Try again in a minute."
    elif "empty list" in low or "blank" in low:
        category = "blank_document"
        user_msg = "Document looks blank or unreadable -- nothing to extract."
    elif "vision parsing error" in low or "llm parsing error" in low:
        category = "llm_error"

    return {
        "document_type": None,
        "vendor": None,
        "dates": {"document_date": None, "due_date": None},
        "money": {
            "total_amount": None,
            "tax_amount": None,
            "net_amount": None,
            "currency": "EUR",
        },
        "line_items": [],
        "ledger_codes": [],
        "needs_manual_input": True,
        "parser_failure_category": category,
        "warnings": [user_msg],
    }
