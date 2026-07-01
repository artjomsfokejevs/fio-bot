"""Tests for services/bookkeeper_report — XLSX + PDF expense reports.

Seeds a handful of documents with different payment_method + ledger_code
combos, then asserts:
  · build_report_data() groups them by source → ledger → per-currency
    subtotals + grand totals
  · generate_xlsx() returns non-empty XLSX bytes with header + row cells
  · generate_pdf() returns a valid PDF byte string
"""
from __future__ import annotations

import io
import json
from datetime import datetime

import pytest

from services import db, bookkeeper_report as bk


@pytest.fixture(autouse=True)
def _clean():
    db.init_db()
    conn = db.get_connection()
    try:
        conn.execute("DELETE FROM documents WHERE id LIKE 'pytest_bk_%'")
        conn.commit()
    finally:
        conn.close()
    yield


def _seed(doc_id, *, amount, ledger, method="bank_transfer",
          vendor="Test Vendor", currency="EUR", period="2026-06",
          net=None, vat=None, legal_entity="ALPS2ALPS_OU",
          payment_executed_at="2026-06-15T10:00:00"):
    if net is None: net = round(amount * 0.83, 2)
    if vat is None: vat = round(amount - net, 2)
    parsed = json.dumps({
        "money": {"total_amount": amount, "net_amount": net,
                   "tax_amount": vat, "currency": currency},
        "invoice": {"number": "INV-" + doc_id[-6:]},
    })
    conn = db.get_connection()
    try:
        conn.execute(
            "INSERT INTO documents (id, filename, file_size, uploaded_at, "
            " uploaded_by, status, amount, currency, currency_orig, "
            " profit_center, ledger_code, period, vendor, legal_entity, "
            " payment_method, payment_executed_at, parsed_json) "
            "VALUES (?, 'x.pdf', 0, '2026-06-14T10:00:00', 'Pytest User', "
            " 'paid', ?, ?, ?, 'AA', ?, ?, ?, ?, ?, ?, ?)",
            (doc_id, amount, currency, currency, ledger, period,
             vendor, legal_entity, method, payment_executed_at, parsed),
        )
        conn.commit()
    finally:
        conn.close()


def test_build_report_groups_by_source_and_ledger():
    _seed("pytest_bk_1", amount=98.60, ledger="4_1_Degviela",
          method="bank_transfer", vendor="Circle K Latvia SIA")
    _seed("pytest_bk_2", amount=7.26, ledger="4_1_Degviela",
          method="card", vendor="CostPocket")
    _seed("pytest_bk_3", amount=6.90, ledger="4_3_Taksi",
          method="card", vendor="MODULAN SIA")
    data = bk.build_report_data(period="2026-06", legal_entity="ALPS2ALPS_OU")
    assert data["period_label"] == "June 2026"
    assert data["row_count"] == 3
    # All 3 payment methods (bank_transfer + card) fall under one
    # bucket ("Company funds"), and are split by ledger inside.
    sources = data["sources"]
    assert len(sources) == 1
    src = sources[0]
    assert src["label"] == "Company funds"
    ledgers = {g["ledger_code"] for g in src["groups"]}
    assert ledgers == {"4_1_Degviela", "4_3_Taksi"}


def test_build_report_subtotals_and_grand_totals():
    _seed("pytest_bk_a", amount=100.00, net=80.00, vat=20.00,
          ledger="4_1_X", currency="EUR")
    _seed("pytest_bk_b", amount=50.00,  net=40.00, vat=10.00,
          ledger="4_1_X", currency="EUR")
    _seed("pytest_bk_c", amount=30.00,  net=30.00, vat=0.00,
          ledger="4_2_Y", currency="USD")
    data = bk.build_report_data(period="2026-06", legal_entity="ALPS2ALPS_OU")
    grp_x = next(g for g in data["sources"][0]["groups"]
                 if g["ledger_code"] == "4_1_X")
    subs_x = {s["currency"]: s for s in grp_x["subtotals"]}
    assert subs_x["EUR"]["net"] == pytest.approx(120.0)
    assert subs_x["EUR"]["vat"] == pytest.approx(30.0)
    assert subs_x["EUR"]["total"] == pytest.approx(150.0)
    grand = data["grand_totals_by_source_and_currency"]["Company funds"]
    assert grand["EUR"]["total"] == pytest.approx(150.0)
    assert grand["USD"]["total"] == pytest.approx(30.0)


def test_build_report_row_view_shapes():
    _seed("pytest_bk_r", amount=98.60, net=81.49, vat=17.11,
          ledger="4_1", vendor="Circle K Latvia SIA",
          payment_executed_at="2026-05-31T09:53:00",
          method="bank_transfer")
    data = bk.build_report_data(period="2026-06", legal_entity="ALPS2ALPS_OU")
    row = data["sources"][0]["groups"][0]["rows"][0]
    assert row["n"] == 1
    assert row["issue_date"] == "31/05/2026"
    assert row["tx_method"] == "bank transfer"
    assert row["vendor"] == "Circle K Latvia SIA"
    assert row["net"] == pytest.approx(81.49)
    assert row["vat"] == pytest.approx(17.11)
    assert row["total"] == pytest.approx(98.60)
    assert row["currency"] == "EUR"


def test_generate_xlsx_produces_valid_workbook():
    _seed("pytest_bk_x1", amount=100.0, ledger="4_1_X")
    data = bk.build_report_data(period="2026-06", legal_entity="ALPS2ALPS_OU")
    xlsx = bk.generate_xlsx(data)
    assert isinstance(xlsx, bytes)
    assert xlsx.startswith(b"PK")  # openpyxl → real ZIP
    # Round-trip
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    assert ws["A1"].value == "Expense Report"
    # Header row of the first data table should live somewhere on the sheet
    found_hdr = False
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "#" and "Issued" in row and "Total" in row:
            found_hdr = True
            break
    assert found_hdr, "Expense Report table header row not found"

    # All visible strings in the sheet must be ASCII / English — no
    # residual Latvian tokens like "Izdevumu" or "Jūnijs".
    forbidden = ("Izdevumu", "atskaite", "Jūnijs", "Uzņēmuma", "Starpsumma",
                  "Piegādātājs", "pārskaitījumu", "Maksāts")
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str):
                for tok in forbidden:
                    assert tok not in cell, f"Latvian token '{tok}' leaked into XLSX cell: {cell!r}"


def test_generate_pdf_produces_valid_pdf():
    _seed("pytest_bk_p1", amount=200.0, ledger="4_2_Y", currency="EUR")
    _seed("pytest_bk_p2", amount=50.0,  ledger="4_2_Y", currency="EUR")
    data = bk.build_report_data(period="2026-06", legal_entity="ALPS2ALPS_OU")
    pdf = bk.generate_pdf(data)
    assert isinstance(pdf, bytes)
    assert pdf.startswith(b"%PDF-"), "generated bytes are not a PDF"
    # At least a few KB — cheap sanity that content actually landed
    assert len(pdf) > 1500


def test_build_bad_period_is_safe():
    # Malformed period must not crash — it should either raise ValueError or
    # return an empty report; either outcome is safe for the export routes.
    try:
        data = bk.build_report_data(period="2026/06")
    except ValueError:
        return
    assert data["row_count"] == 0


def test_empty_period_returns_empty_report_not_error():
    data = bk.build_report_data(period="1999-01", legal_entity="ALPS2ALPS_OU")
    assert data["row_count"] == 0
    assert data["sources"] == []
